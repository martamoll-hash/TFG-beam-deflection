
import cv2
import numpy as np
import matplotlib.pyplot as plt
import os
import sys
import datetime
from enum import Enum

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import (Font, PatternFill, Alignment,
                                  Border, Side, numbers)
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    print("[WARNING] openpyxl not found — Excel export disabled. "
          "Install with: pip install openpyxl")

# ═══════════════════════════════════════════════════════════════════════════════
# USER CONFIGURATION — edit these values before running
# ═══════════════════════════════════════════════════════════════════════════════

CAMERA_INDEX = 0          # OpenCV camera index

# Beam cross-section and material (aluminium default)
b_mm     = 20.0          # width  [mm]
h_mm     =  2.0           # height [mm]
E_MPa    = 69000.0        # Young's modulus [MPa = N/mm²]
rho_kgm3 = 2700.0         # density [kg/m³]

# Beam span — used to lock the horizontal calibration to a known distance
L_KNOWN_MM = 600.0        # total beam length [mm]

# Grid display
GRID_PX_DEFAULT = 80      # initial grid step [px] before calibration

# Auto-contour detection parameters (Canny)
AUTO_CANNY_LOW  = 20
AUTO_CANNY_HIGH = 80
AUTO_BAND_HALF  = 120     # px band around neutral axis searched for edges
AUTO_MEDFILT_K  = 11      # median filter kernel for smoothing detected edge

# ═══════════════════════════════════════════════════════════════════════════════
# DERIVED CONSTANTS
# ═══════════════════════════════════════════════════════════════════════════════

I_mm4    = b_mm * h_mm**3 / 12.0
EI       = E_MPa * I_mm4
A_mm2    = b_mm * h_mm
rho_Nmm3 = rho_kgm3 * 9.81e-9     # [N/mm³]
w_Nmm    = rho_Nmm3 * A_mm2        # self-weight per unit length [N/mm]

WIN        = "Beam Deflection Analysis"
ZOOM_LEVELS = [1.0, 1.5, 2.0, 3.0, 4.0, 6.0, 8.0]

print(f"[INFO] b={b_mm} mm  h={h_mm} mm  I={I_mm4:.4f} mm⁴  EI={EI:.0f} N·mm²")
print(f"[INFO] A={A_mm2:.1f} mm²  w={w_Nmm*1000:.4f} mN/mm")
print(f"[INFO] Self-weight: empirical two-digitisation subtraction")


# ═══════════════════════════════════════════════════════════════════════════════
# BEAM CASE ENUM
# ═══════════════════════════════════════════════════════════════════════════════

class BeamCase(Enum):
    CANTILEVER_FREE_END      = 1   # cantilever, load at x = L
    CANTILEVER_INTERMEDIATE  = 2   # cantilever, load at x = a < L
    SIMPLY_SUPPORTED         = 3   # simply supported, load at x = a
    OVERHANG_RIGHT_FREE_END  = 4   # pin A, roller B, load at free end C
    OVERHANG_LOAD_BETWEEN    = 5   # pin A, load at B, roller C, free end D
    OVERHANG_LOAD_IN_SPAN    = 6   # pin A, roller B, load at C, free end D


# ═══════════════════════════════════════════════════════════════════════════════
# PHYSICS — EULER-BERNOULLI ANALYTICAL SOLUTIONS
# ═══════════════════════════════════════════════════════════════════════════════

def v_self_weight_analytical(x, L):
    """
    Exact tip deflection due to self-weight for a cantilever (fixed at x=0).
    v(x) = w/(24EI) · (6L²x² − 4Lx³ + x⁴)

    Kept for reference / printout only. NOT used in the fitting pipeline.
    The empirical two-digitisation strategy replaces this.
    """
    return w_Nmm / (24.0 * EI) * (6*L**2*x**2 - 4*L*x**3 + x**4)


# ─────────────────────────────────────────────────────────────────────────────
# Case 1: Cantilever, load P at free end (x = L)
# ─────────────────────────────────────────────────────────────────────────────
# v(x) = P/(6EI) · (3Lx² - x³)        for 0 ≤ x ≤ L
# M(x) = P·(L - x)                     (from equilibrium, not differentiation)
# V(x) = P                             (constant throughout)

def fit_cantilever_free_end(x_mm, v_mm, L):
    """
    Least-squares fit of P using exact Euler-Bernoulli shape function.
    Boundary conditions v(0)=0 and v'(0)=0 are satisfied analytically.
    Returns (P_N, v_func, M_func, V_func) or (None,...) on failure.
    """
    phi   = (3*L*x_mm**2 - x_mm**3) / (6.0*EI)
    denom = float(np.dot(phi, phi))
    if denom < 1e-30:
        print("[ERROR] Degenerate fit — check digitised points.")
        return None, None, None, None
    P = float(np.dot(phi, v_mm) / denom)

    def v_func(x):
        return P / (6.0*EI) * (3*L*x**2 - x**3)

    def M_func(x):
        return -P * (L - x)

    def V_func(x):
        return np.full_like(np.asarray(x, dtype=float), P)

    return P, v_func, M_func, V_func


# ─────────────────────────────────────────────────────────────────────────────
# Case 2: Cantilever, load P at intermediate position x = a  (a < L)
# ─────────────────────────────────────────────────────────────────────────────
# Region 1 (0 ≤ x ≤ a):  v₁(x) = P/(6EI)·(3ax² - x³)
# Region 2 (a < x ≤ L):  v₂(x) = Pa²/(6EI)·(3x - a)      [zero curvature]
# M(x) = P·(a-x)  for x≤a ;  0  for x>a
# V(x) = P        for x≤a ;  0  for x>a

def fit_cantilever_intermediate(x_mm, v_mm, L, a):
    """
    Two-segment cantilever with load at x=a.
    'a' is provided by the user (load position in mm from fixed end).
    Least-squares fit for P using piecewise exact shape function.
    """
    def phi(x):
        p = np.zeros_like(x, dtype=float)
        m1 = x <= a
        m2 = ~m1
        p[m1] = (3*a*x[m1]**2 - x[m1]**3) / (6.0*EI)
        p[m2] = a**2 * (3*x[m2] - a)       / (6.0*EI)
        return p

    ph    = phi(x_mm)
    denom = float(np.dot(ph, ph))
    if denom < 1e-30:
        print("[ERROR] Degenerate fit — check digitised points.")
        return None, None, None, None
    P = float(np.dot(ph, v_mm) / denom)

    def v_func(x):
        x  = np.asarray(x, dtype=float)
        vv = np.zeros_like(x)
        m1 = x <= a;  m2 = ~m1
        vv[m1] = P/(6.0*EI) * (3*a*x[m1]**2 - x[m1]**3)
        vv[m2] = P*a**2/(6.0*EI) * (3*x[m2] - a)
        return vv

    def M_func(x):
        x  = np.asarray(x, dtype=float)
        mm = np.zeros_like(x)
        m1 = x <= a
        mm[m1] = -P * (a - x[m1])
        return mm

    def V_func(x):
        x  = np.asarray(x, dtype=float)
        vv = np.zeros_like(x)
        vv[x <= a] = P
        return vv

    return P, v_func, M_func, V_func


# ─────────────────────────────────────────────────────────────────────────────
# Case 3: Simply supported beam, load P at position x = a
# ─────────────────────────────────────────────────────────────────────────────
# Reactions: Ra = P·b/L   Rb = P·a/L   where b = L - a
# Region 1 (0 ≤ x ≤ a):
#   v₁(x) = P·b/(6EI·L) · (L²x - b²x - x³)
# Region 2 (a < x ≤ L):
#   v₂(x) = P·a/(6EI·L) · (L²(x-a) - (x-a)³ + ... )
#   — written using symmetry: v₂(x) = v₁(L-x) evaluated for load at (L-a)
# M(x) = Ra·x            for x ≤ a
# M(x) = Ra·x - P·(x-a) for x > a
# V(x) = Ra              for x ≤ a ;  V(x) = Ra - P  for x > a

def fit_simply_supported(x_mm, v_mm, L, a):
    """
    Simply supported beam with point load at x=a.
    Least-squares fit for P. BCs v(0)=0, v(L)=0 are satisfied analytically.
    """
    b = L - a

    def phi(x):
        p  = np.zeros_like(x, dtype=float)
        m1 = x <= a
        m2 = ~m1
        # Region 1
        p[m1] = b / (6.0*EI*L) * (L**2*x[m1] - b**2*x[m1] - x[m1]**3)
        # Region 2 — exact Euler-Bernoulli solution second segment
        xr     = x[m2]
        p[m2]  = a / (6.0*EI*L) * (L**2*(xr-a) - (xr-a)**3 - a**2*(xr-a) + (L**2-b**2-a**2)*0)
        # Cleaner form for region 2:
        # v₂ = Pa(L-x)/(6EIL)·(2Lx - x² - a²)  — standard formula
        p[m2]  = P_ss_region2_phi(xr, a, b, L)
        return p

    def P_ss_region2_phi(xr, a, b, L):
        # φ₂ = a(L-x)/(6EIL)·(2Lx - x² - a²)
        return a*(L-xr)/(6.0*EI*L) * (2*L*xr - xr**2 - a**2)

    # Build φ properly (region 2 formula needs P removed — it's just the shape)
    ph = np.zeros_like(x_mm, dtype=float)
    m1 = x_mm <= a;  m2 = ~m1
    ph[m1] = b/(6.0*EI*L) * (L**2*x_mm[m1] - b**2*x_mm[m1] - x_mm[m1]**3)
    ph[m2] = a*(L-x_mm[m2])/(6.0*EI*L) * (2*L*x_mm[m2] - x_mm[m2]**2 - a**2)

    denom = float(np.dot(ph, ph))
    if denom < 1e-30:
        print("[ERROR] Degenerate fit — check digitised points.")
        return None, None, None, None
    P = float(np.dot(ph, v_mm) / denom)

    Ra = P * b / L
    Rb = P * a / L

    def v_func(x):
        x  = np.asarray(x, dtype=float)
        vv = np.zeros_like(x)
        m1 = x <= a;  m2 = ~m1
        vv[m1] = P*b/(6.0*EI*L) * (L**2*x[m1] - b**2*x[m1] - x[m1]**3)
        vv[m2] = P*a/(6.0*EI*L) * (L**2*(x[m2]-a) - (x[m2]-a)**3 + \
                 (L**2 - a**2 - (L-x[m2])**2) * 0)
        # Standard second-segment formula:
        vv[m2] = P*a*(L-x[m2])/(6.0*EI*L) * (2*L*x[m2] - x[m2]**2 - a**2)
        return vv

    def M_func(x):
        x  = np.asarray(x, dtype=float)
        mm = np.zeros_like(x)
        m1 = x <= a;  m2 = ~m1
        mm[m1] = Ra * x[m1]
        mm[m2] = Ra * x[m2] - P * (x[m2] - a)
        return mm

    def V_func(x):
        x  = np.asarray(x, dtype=float)
        vv = np.zeros_like(x)
        m1 = x <= a;  m2 = ~m1
        vv[m1] =  Ra
        vv[m2] =  Ra - P
        return vv

    return P, v_func, M_func, V_func


# ─────────────────────────────────────────────────────────────────────────────
# Case 4: Overhanging beam — load P at right free end C
# ─────────────────────────────────────────────────────────────────────────────
# Geometry:  A ——— B ——— C
#            pin  roller  free end (load P)
#            |<La>|< Lb >|
# Origin x=0 at A.
#
# Reactions: RB = PL/La  (upward)   RA = -PLb/La  (downward)
#            where L = La + Lb
#
# M(x):  Span AB: M = -(PLb/La)·x
#        Span BC: M = -P(L-x)
#
# v(x):  Span AB: v = -PLb·x·(x²-La²) / (6·EI·La)
#        Span BC: v = P(x-La)·(x²-2Lx+La²+LaLb-2Lb²+...) / 6EI

def fit_overhang_right_free_end(x_mm, dv_mm, La, Lb):
    """
    Least-squares fit for P. x in [0, La+Lb], origin at pin A.
    phi(x) = v(x)/P  evaluated piecewise from the analytical solution.
    """
    L  = La + Lb
    x  = np.asarray(x_mm, dtype=float)
    ph = np.zeros_like(x)
    m1 = x <= La;  m2 = ~m1

    # Span AB: v_a(x) = -Lb·x·(x²-La²) / (6EI·La)
    ph[m1] = -Lb * x[m1] * (x[m1]**2 - La**2) / (6.0*EI*La)

    # Span BC: v_b(x) = (x-La)·(x²-2Lx+La²+LaLb) / (6EI)
    # Derived from symbolic solution: factor(vb) = (x-La)*(x²-(2La+3Lb)*x + La²+LaLb-2Lb² +...)
    # Full: vb = P/6EI * (x-La)*(x^2 - (2*La + 3*Lb)*x + La^2 + La*Lb - 2*Lb^2 + 2*Lb^2)
    # Simplification confirmed by sympy: vb = -Lb^2*P*(La+Lb)/3 at x=La+Lb  ✓
    xi = x[m2]
    ph[m2] = (xi - La) * (xi**2 - (2*La + 3*Lb)*xi + La**2 + La*Lb) / (6.0*EI)

    denom = float(np.dot(ph, ph))
    if denom < 1e-30:
        print("[ERROR] Degenerate fit."); return None, None, None, None
    P_signed = float(np.dot(ph, dv_mm) / denom)
    P        = abs(P_signed)
    sign     = 1.0 if P_signed >= 0 else -1.0   # deflection direction from image

    RA = -P*Lb/La
    RB =  P*L/La

    def v_func(x):
        x  = np.asarray(x, dtype=float)
        vv = np.zeros_like(x)
        m1 = x <= La;  m2 = ~m1
        vv[m1] = -P*Lb*x[m1]*(x[m1]**2 - La**2) / (6.0*EI*La)
        xi = x[m2]
        vv[m2] = P*(xi-La)*(xi**2 - (2*La + 3*Lb)*xi + La**2 + La*Lb) / (6.0*EI)
        return vv * sign   # respect measured deflection direction

    def M_func(x):
        x  = np.asarray(x, dtype=float)
        mm = np.zeros_like(x)
        m1 = x <= La;  m2 = ~m1
        mm[m1] = -(P*Lb/La) * x[m1]
        mm[m2] = -P*(L - x[m2])
        return mm

    def V_func(x):
        x  = np.asarray(x, dtype=float)
        vv = np.zeros_like(x)
        m1 = x <= La;  m2 = ~m1
        vv[m1] = RA
        vv[m2] = P
        return vv

    return P, v_func, M_func, V_func


# ─────────────────────────────────────────────────────────────────────────────
# Case 5: Overhanging — load P at B (between supports), free end D at right
# ─────────────────────────────────────────────────────────────────────────────
# Geometry:  A ——— B ——— C ——— D
#            pin  load  roller  free
#            |<La>|< Lb>|< Lc >|
# Origin x=0 at A.
#
# Reactions: RC = PLa/(La+Lb)   RA = PLb/(La+Lb)
# M(x):  AB: M = PLb/(La+Lb)·x
#        BC: M = PLa/(La+Lb)·(La+Lb-x)
#        CD: M = 0  (no load, free end)

def fit_overhang_load_between(x_mm, dv_mm, La, Lb, Lc):
    """
    Least-squares fit for P. Origin at pin A.
    Three segments: AB (0..La), BC (La..La+Lb), CD (La+Lb..La+Lb+Lc).
    """
    Sab = La + Lb   # position of roller C
    L   = Sab + Lc  # total length
    x   = np.asarray(x_mm, dtype=float)
    ph  = np.zeros_like(x)

    m1 = x <= La
    m2 = (x > La) & (x <= Sab)
    m3 = x > Sab

    # Span AB: v_a = Lb·x·(x²-La²-2·La·Lb) / (6·EI·(La+Lb))
    ph[m1] = Lb*x[m1]*(x[m1]**2 - La**2 - 2*La*Lb) / (6.0*EI*Sab)

    # Span BC: v_b = -La·(x-La-Lb)·(x²-2(La+Lb)x+La²) / (6·EI·(La+Lb))
    xi = x[m2]
    ph[m2] = -La*(xi - Sab)*(xi**2 - 2*Sab*xi + La**2) / (6.0*EI*Sab)

    # Span CD: v_c = La·Lb·(2La+Lb)·(x-La-Lb) / (6·EI·(La+Lb))
    ph[m3] = La*Lb*(2*La+Lb)*(x[m3]-Sab) / (6.0*EI*Sab)

    denom = float(np.dot(ph, ph))
    if denom < 1e-30:
        print("[ERROR] Degenerate fit."); return None, None, None, None
    P_signed = float(np.dot(ph, dv_mm) / denom)
    P        = abs(P_signed)
    sign     = 1.0 if P_signed >= 0 else -1.0

    RA = P*Lb/Sab
    RC = P*La/Sab

    def v_func(x):
        x  = np.asarray(x, dtype=float)
        vv = np.zeros_like(x)
        m1 = x <= La
        m2 = (x > La) & (x <= Sab)
        m3 = x > Sab
        vv[m1] = P*Lb*x[m1]*(x[m1]**2 - La**2 - 2*La*Lb) / (6.0*EI*Sab)
        xi = x[m2]
        vv[m2] = -P*La*(xi-Sab)*(xi**2 - 2*Sab*xi + La**2) / (6.0*EI*Sab)
        vv[m3] = P*La*Lb*(2*La+Lb)*(x[m3]-Sab) / (6.0*EI*Sab)
        return vv * sign

    def M_func(x):
        x  = np.asarray(x, dtype=float)
        mm = np.zeros_like(x)
        m1 = x <= La
        m2 = (x > La) & (x <= Sab)
        # CD: M=0
        mm[m1] = RA * x[m1]
        mm[m2] = P*La/Sab * (Sab - x[m2])
        return mm

    def V_func(x):
        x  = np.asarray(x, dtype=float)
        vv = np.zeros_like(x)
        m1 = x <= La
        m2 = (x > La) & (x <= Sab)
        m3 = x > Sab
        vv[m1] =  RA
        vv[m2] =  RA - P
        vv[m3] =  0.0
        return vv

    return P, v_func, M_func, V_func


# ─────────────────────────────────────────────────────────────────────────────
# Case 6: Overhanging — load P in right overhang at C
# ─────────────────────────────────────────────────────────────────────────────
# Geometry:  A ——— B ——— C ——— D
#            pin  roller  load  free
#            |<La>|< Lb >|< Lc>|
# Origin x=0 at A.
#
# Reactions: RB = P(La+Lb)/La   RA = -PLb/La
# M(x):  AB: M = -(PLb/La)·x
#        BC: M = P(x-La-Lb)
#        CD: M = 0

def fit_overhang_load_in_span(x_mm, dv_mm, La, Lb, Lc):
    """
    Least-squares fit for P. Origin at pin A.
    Three segments: AB (0..La), BC (La..La+Lb), CD (La+Lb..La+Lb+Lc).
    """
    Sab = La + Lb   # position of load C
    L   = Sab + Lc
    x   = np.asarray(x_mm, dtype=float)
    ph  = np.zeros_like(x)

    m1 = x <= La
    m2 = (x > La) & (x <= Sab)
    m3 = x > Sab

    # Span AB: v_a = -Lb·x·(x-La)·(x+La) / (6·EI·La)
    ph[m1] = -Lb*x[m1]*(x[m1]-La)*(x[m1]+La) / (6.0*EI*La)

    # Span BC: v_b = (x-La)·(x²-2(La+Lb)x+La²+LaLb-2Lb²+...)/6EI  ← from case 3 derivation
    # Sympy result: vb = P/6EI*(x-La)*(x^2-2*(La+Lb)*x+La^2+La*Lb-2*Lb^2+2*Lb^2)
    # Confirmed: vb = P*(x-La)*(x^2 - 2*Sab*x + La^2) / (6EI)  ... let's use sympy-verified form
    xi = x[m2]
    ph[m2] = (xi-La)*(xi**2 - (2*La+3*Lb)*xi + La**2 + La*Lb) / (6.0*EI)
    
    # Span CD: v_c = -Lb·(2La+3Lb)·(x-La-Lb) / (6EI)  +  Lb·(La+Lb)·(2La+Lb)/(6EI)
    # Linear in x. phi = v/P:
    ph[m3] = -Lb*(-2*La**2 - 3*La*Lb + 2*La*x[m3] - Lb**2 + 3*Lb*x[m3]) / (6.0*EI)

    denom = float(np.dot(ph, ph))
    if denom < 1e-30:
        print("[ERROR] Degenerate fit."); return None, None, None, None
    P_signed = float(np.dot(ph, dv_mm) / denom)
    P        = abs(P_signed)
    sign     = 1.0 if P_signed >= 0 else -1.0

    RA = -P*Lb/La
    RB =  P*(La+Lb)/La

    def v_func(x):
        x  = np.asarray(x, dtype=float)
        vv = np.zeros_like(x)
        m1 = x <= La
        m2 = (x > La) & (x <= Sab)
        m3 = x > Sab
        vv[m1] = -P*Lb*x[m1]*(x[m1]-La)*(x[m1]+La) / (6.0*EI*La)
        xi = x[m2]
        vv[m2] = P*(xi-La)*(xi**2 - (2*La+3*Lb)*xi + La**2 + La*Lb) / (6.0*EI)
        vv[m3] = -P*Lb*(-2*La**2 - 3*La*Lb + 2*La*x[m3] - Lb**2 + 3*Lb*x[m3]) / (6.0*EI)
        return vv * sign

    def M_func(x):
        x  = np.asarray(x, dtype=float)
        mm = np.zeros_like(x)
        m1 = x <= La
        m2 = (x > La) & (x <= Sab)
        # CD: M=0
        mm[m1] = RA * x[m1]
        mm[m2] = P*(x[m2] - Sab)
        return mm

    def V_func(x):
        x  = np.asarray(x, dtype=float)
        vv = np.zeros_like(x)
        m1 = x <= La
        m2 = (x > La) & (x <= Sab)
        m3 = x > Sab
        vv[m1] = RA          # -PLb/La
        vv[m2] = RA + RB     # -PLb/La + P(La+Lb)/La = P
        vv[m3] = 0.0
        return vv

    return P, v_func, M_func, V_func


# ═══════════════════════════════════════════════════════════════════════════════
# GRID CLASS
# ═══════════════════════════════════════════════════════════════════════════════

class Grid:
    """
    Real-time mm grid overlay.

    Before calibration: fixed pixel step (GRID_PX_DEFAULT).
    After calibration:  step snapped to a round mm value using the measured
                        px_per_mm scale. If separate X and Y scales are
                        provided (anisotropic case), both axes are drawn
                        with their own step in pixels but labelled in mm.

    Origin (ox, oy) in display pixels is movable via middle-button drag.

    Keyboard shortcuts (active in all phases):
      G       — toggle visibility
      + / =   — increase grid step
      -       — decrease grid step
      Middle-drag — pan origin
    """

    NICE_MM = [5, 10, 20, 25, 50, 100, 150, 200]

    def __init__(self, dw, dh):
        self.dw        = dw
        self.dh        = dh
        self.ox        = dw // 2
        self.oy        = dh // 2
        self.step_px   = GRID_PX_DEFAULT
        self.step_px_y = GRID_PX_DEFAULT
        self.step_mm   = None
        self.px_per_mm   = None   # X scale [px/mm] in display coords
        self.px_per_mm_y = None   # Y scale [px/mm] in display coords
        self.visible   = True
        self._drag     = False
        self._dx = self._dy = 0

    def calibrate(self, px_per_mm_x_disp, px_per_mm_y_disp=None):
        """Set scale(s) and snap grid step to nearest nice mm value."""
        self.px_per_mm   = px_per_mm_x_disp
        self.px_per_mm_y = px_per_mm_y_disp or px_per_mm_x_disp
        best = min(self.NICE_MM,
                   key=lambda m: abs(m * px_per_mm_x_disp - GRID_PX_DEFAULT))
        self.step_mm   = best
        self.step_px   = int(round(best * px_per_mm_x_disp))
        self.step_px_y = int(round(best * self.px_per_mm_y))
        print(f"[GRID] X: {self.step_mm} mm = {self.step_px} px  |  "
              f"Y: {self.step_mm} mm = {self.step_px_y} px")

    def set_origin(self, ox, oy):
        self.ox = int(np.clip(ox, 0, self.dw))
        self.oy = int(np.clip(oy, 0, self.dh))

    def increase_step(self):
        if self.px_per_mm and self.step_mm:
            idx = self.NICE_MM.index(self.step_mm) \
                  if self.step_mm in self.NICE_MM else 0
            idx = min(idx + 1, len(self.NICE_MM) - 1)
            self.step_mm   = self.NICE_MM[idx]
            self.step_px   = int(round(self.step_mm * self.px_per_mm))
            self.step_px_y = int(round(self.step_mm * (self.px_per_mm_y
                                                        or self.px_per_mm)))
        else:
            self.step_px   = min(self.step_px + 20, 400)
            self.step_px_y = self.step_px

    def decrease_step(self):
        if self.px_per_mm and self.step_mm:
            idx = self.NICE_MM.index(self.step_mm) \
                  if self.step_mm in self.NICE_MM else 0
            idx = max(idx - 1, 0)
            self.step_mm   = self.NICE_MM[idx]
            self.step_px   = max(int(round(self.step_mm * self.px_per_mm)), 2)
            self.step_px_y = max(int(round(self.step_mm * (self.px_per_mm_y
                                                            or self.px_per_mm))), 2)
        else:
            self.step_px   = max(self.step_px - 20, 20)
            self.step_px_y = self.step_px

    def draw(self, img):
        """Draw grid and axes on img in-place."""
        if not self.visible:
            return
        h, w     = img.shape[:2]
        ox, oy   = self.ox, self.oy
        step_x   = max(self.step_px,   2)
        step_y   = max(self.step_px_y, 2)
        cal      = self.px_per_mm is not None

        COL_GRID  = (60,  60,  60)
        COL_AXIS_X = (0,  200, 255)   # cyan  — horizontal axis
        COL_AXIS_Y = (255,180,  0)    # amber — vertical axis
        COL_LABEL  = (220,220, 220)

        # Vertical lines (X axis)
        x = ox; vx = 0
        while x < w:
            color = COL_AXIS_Y if x == ox else COL_GRID
            thick = 2           if x == ox else 1
            cv2.line(img, (x, 0), (x, h), color, thick)
            if cal and x != ox:
                cv2.putText(img, f"{int(vx)}", (x+2, oy+18),
                            cv2.FONT_HERSHEY_SIMPLEX, 0.36, COL_AXIS_X, 1)
            x += step_x; vx += (self.step_mm or step_x)

        x = ox - step_x; vx = -(self.step_mm or step_x)
        while x > 0:
            cv2.line(img, (x, 0), (x, h), COL_GRID, 1)
            if cal:
                cv2.putText(img, f"{int(vx)}", (x+2, oy+18),
                            cv2.FONT_HERSHEY_SIMPLEX, 0.36, COL_AXIS_X, 1)
            x -= step_x; vx -= (self.step_mm or step_x)

        # Horizontal lines (Y axis)
        y = oy; vy = 0
        while y < h:
            color = COL_AXIS_X if y == oy else COL_GRID
            thick = 2           if y == oy else 1
            cv2.line(img, (0, y), (w, y), color, thick)
            if cal and y != oy:
                cv2.putText(img, f"{int(vy)}", (ox+4, y-3),
                            cv2.FONT_HERSHEY_SIMPLEX, 0.36, COL_AXIS_Y, 1)
            y += step_y; vy += (self.step_mm or step_y)

        y = oy - step_y; vy = -(self.step_mm or step_y)
        while y > 0:
            cv2.line(img, (0, y), (w, y), COL_GRID, 1)
            if cal:
                cv2.putText(img, f"{int(vy)}", (ox+4, y-3),
                            cv2.FONT_HERSHEY_SIMPLEX, 0.36, COL_AXIS_Y, 1)
            y -= step_y; vy -= (self.step_mm or step_y)

        # Labels
        if cal:
            lbl = f"x [mm]  step={self.step_mm} mm"
            if self.px_per_mm_y and abs(self.px_per_mm_y - self.px_per_mm) > 0.01:
                lbl += "  [anisotropic]"
            cv2.putText(img, lbl,
                        (w-220, oy-6), cv2.FONT_HERSHEY_SIMPLEX, 0.36, COL_AXIS_X, 1)
            cv2.putText(img, "v [mm]",
                        (ox+6, 16),    cv2.FONT_HERSHEY_SIMPLEX, 0.36, COL_AXIS_Y, 1)
        else:
            cv2.putText(img, f"Grid: {step_x} px/cell (uncalibrated)",
                        (10, h-10),    cv2.FONT_HERSHEY_SIMPLEX, 0.36, COL_LABEL, 1)

        cv2.circle(img, (ox, oy), 5, (255,255,255), -1)
        cv2.putText(img, "0,0" if cal else "orig",
                    (ox+6, oy-6), cv2.FONT_HERSHEY_SIMPLEX, 0.36, (255,255,255), 1)

    def handle_key(self, k):
        if k in (ord('g'), ord('G')):
            self.visible = not self.visible; return True
        if k in (ord('+'), ord('=')):
            self.increase_step(); return True
        if k == ord('-'):
            self.decrease_step(); return True
        return False

    def handle_mouse_drag(self, event, x, y, flags):
        if event == cv2.EVENT_MBUTTONDOWN:
            self._drag = True; self._dx = x; self._dy = y
        elif event == cv2.EVENT_MBUTTONUP:
            self._drag = False
        elif event == cv2.EVENT_MOUSEMOVE and self._drag:
            self.ox = int(np.clip(self.ox + (x - self._dx), 0, self.dw))
            self.oy = int(np.clip(self.oy + (y - self._dy), 0, self.dh))
            self._dx = x; self._dy = y
            return True
        return False


# ═══════════════════════════════════════════════════════════════════════════════
# ZOOM VIEW
# ═══════════════════════════════════════════════════════════════════════════════

class ZoomView:
    def __init__(self, frame_orig, scale):
        self.orig  = frame_orig
        self.scale = scale
        self.H, self.W = frame_orig.shape[:2]
        self.zoom  = 1.0
        self.cx    = self.W // 2
        self.cy    = self.H // 2

    def get_roi(self):
        z  = self.zoom
        dw = int(self.W / z / 2); dh = int(self.H / z / 2)
        x0 = int(np.clip(self.cx - dw, 0, self.W - 2*dw))
        y0 = int(np.clip(self.cy - dh, 0, self.H - 2*dh))
        return x0, y0, x0+2*dw, y0+2*dh

    def get_disp_size(self):
        return int(self.W*self.scale), int(self.H*self.scale)

    def render(self):
        x0, y0, x1, y1 = self.get_roi()
        dw, dh = self.get_disp_size()
        return cv2.resize(self.orig[y0:y1, x0:x1], (dw, dh))

    def disp_to_orig(self, xd, yd):
        x0, y0, x1, y1 = self.get_roi()
        dw, dh = self.get_disp_size()
        return int(x0 + xd/dw*(x1-x0)), int(y0 + yd/dh*(y1-y0))

    def orig_to_disp(self, xo, yo):
        x0, y0, x1, y1 = self.get_roi()
        dw, dh = self.get_disp_size()
        return int((xo-x0)/(x1-x0)*dw), int((yo-y0)/(y1-y0)*dh)

    def wheel_zoom(self, flags, xd=None, yd=None):
        if xd is not None and yd is not None:
            self.cx, self.cy = self.disp_to_orig(xd, yd)   # zoom toward the cursor
        idx = self._zi()
        self.zoom = ZOOM_LEVELS[int(np.clip(idx+(1 if flags>0 else -1),
                                            0, len(ZOOM_LEVELS)-1))]

    def key_zoom(self, d):
        self.zoom = ZOOM_LEVELS[int(np.clip(self._zi()+d,
                                            0, len(ZOOM_LEVELS)-1))]

    def _zi(self):
        try:    return ZOOM_LEVELS.index(self.zoom)
        except: return 0


# ═══════════════════════════════════════════════════════════════════════════════
# CAMERA
# ═══════════════════════════════════════════════════════════════════════════════

def open_camera(index):
    cap = cv2.VideoCapture(index, cv2.CAP_DSHOW)
    if not cap.isOpened():
        cap = cv2.VideoCapture(index)
    if not cap.isOpened():
        print(f"[ERROR] Cannot open camera {index}."); sys.exit(1)
    cap.set(cv2.CAP_PROP_FRAME_WIDTH,  1280)
    cap.set(cv2.CAP_PROP_FRAME_HEIGHT,  720)
    return cap


def get_scale(cap):
    for _ in range(10):
        ret, frame = cap.read()
        if ret:
            h, w = frame.shape[:2]
            return min(1.0, 1280/w, 720/h), frame
    print("[ERROR] Cannot read from camera."); sys.exit(1)


def freeze_frame(cap, scale, msg, grid=None):
    """Live preview with optional grid. Press SPACE to freeze."""
    print(f"\n  {msg}")
    while True:
        ret, frame = cap.read()
        if not ret: continue
        dw = int(frame.shape[1]*scale); dh = int(frame.shape[0]*scale)
        disp = cv2.resize(frame, (dw, dh))
        if grid: grid.draw(disp)
        cv2.putText(disp, msg, (10,30),
                    cv2.FONT_HERSHEY_SIMPLEX, 0.60, (0,255,255), 2)
        cv2.putText(disp, "SPACE: capture  |  G: grid  |  +/-: step  |  ESC: exit",
                    (10,52), cv2.FONT_HERSHEY_SIMPLEX, 0.40, (0,200,200), 1)
        cv2.imshow(WIN, disp)
        k = cv2.waitKey(30) & 0xFF
        if k == 27: cv2.destroyAllWindows(); sys.exit(0)
        if k == ord(' '): print("  -> Frame captured."); return frame.copy()
        if grid: grid.handle_key(k)


# ═══════════════════════════════════════════════════════════════════════════════
# CLICK HELPERS
# ═══════════════════════════════════════════════════════════════════════════════

def _two_clicks(frame_orig, scale, grid, zv, msg, color, n=2):
    """Collect n left-clicks with zoom. Returns list of (x,y) in original px."""
    state = {'pts': []}

    def redraw():
        img = zv.render()
        grid.draw(img)
        cv2.putText(img, msg, (10,28),
                    cv2.FONT_HERSHEY_SIMPLEX, 0.52, color, 2)
        cv2.putText(img,
            f"Zoom x{zv.zoom:.1f}  |  +/-: zoom  |  Wheel: zoom  |  Z: undo  |  G: grid",
            (10,50), cv2.FONT_HERSHEY_SIMPLEX, 0.38, (200,200,0), 1)
        for i, (xo, yo) in enumerate(state['pts']):
            xd, yd = zv.orig_to_disp(xo, yo)
            cv2.circle(img, (xd, yd), 7, color, -1)
            cv2.putText(img, str(i+1), (xd+9, yd-6),
                        cv2.FONT_HERSHEY_SIMPLEX, 0.5, color, 1)
        if len(state['pts']) == 2:
            xd0, yd0 = zv.orig_to_disp(*state['pts'][0])
            xd1, yd1 = zv.orig_to_disp(*state['pts'][1])
            cv2.line(img, (xd0,yd0), (xd1,yd1), color, 1)
        cv2.imshow(WIN, img)

    def cb(event, x, y, flags, param):
        if grid.handle_mouse_drag(event, x, y, flags): redraw(); return
        if event == cv2.EVENT_MOUSEWHEEL: zv.wheel_zoom(flags, x, y); redraw()
        elif event == cv2.EVENT_LBUTTONDOWN and len(state['pts']) < n:
            state['pts'].append(zv.disp_to_orig(x,y)); redraw()

    cv2.setMouseCallback(WIN, cb)
    redraw()
    while len(state['pts']) < n:
        k = cv2.waitKey(30) & 0xFF
        if k == 27: cv2.destroyAllWindows(); sys.exit(0)
        if k in (ord('z'), ord('Z')) and state['pts']:
            state['pts'].pop(); redraw()
        if k in (ord('+'), ord('=')): zv.key_zoom(+1); redraw()
        if k == ord('-'):              zv.key_zoom(-1); redraw()
        if grid.handle_key(k):        redraw()
    cv2.setMouseCallback(WIN, lambda *a: None)
    return state['pts']


def zoom_one_right_click(frame_orig, scale, msg, grid=None):
    """Single right-click with zoom. Returns (x,y) in original px."""
    zv    = ZoomView(frame_orig, scale)
    state = {'pt': None}

    def redraw():
        img = zv.render()
        if grid: grid.draw(img)
        cv2.putText(img, msg, (10,30),
                    cv2.FONT_HERSHEY_SIMPLEX, 0.58, (255,0,255), 2)
        cv2.putText(img,
            f"Zoom x{zv.zoom:.1f}  |  +/-: zoom  |  Wheel: zoom  |  G: grid",
            (10,52), cv2.FONT_HERSHEY_SIMPLEX, 0.38, (200,200,0), 1)
        if state['pt']:
            xd, yd = zv.orig_to_disp(*state['pt'])
            cv2.circle(img, (xd,yd), 9, (255,0,255), -1)
            cv2.putText(img, "Support", (xd+10, yd-8),
                        cv2.FONT_HERSHEY_SIMPLEX, 0.55, (255,0,255), 2)
        cv2.imshow(WIN, img)

    def cb(event, x, y, flags, param):
        if grid and grid.handle_mouse_drag(event, x, y, flags):
            redraw(); return
        if event == cv2.EVENT_MOUSEWHEEL:
            zv.wheel_zoom(flags, x, y); redraw()
        elif event == cv2.EVENT_RBUTTONDOWN:
            state['pt'] = zv.disp_to_orig(x,y); redraw()

    cv2.setMouseCallback(WIN, cb)
    redraw()
    while state['pt'] is None:
        k = cv2.waitKey(30) & 0xFF
        if k == 27: cv2.destroyAllWindows(); sys.exit(0)
        if k in (ord('z'), ord('Z')): state['pt'] = None; redraw()
        if k in (ord('+'), ord('=')): zv.key_zoom(+1); redraw()
        if k == ord('-'):              zv.key_zoom(-1); redraw()
        if grid and grid.handle_key(k): redraw()
    cv2.setMouseCallback(WIN, lambda *a: None)
    return state['pt']


# ═══════════════════════════════════════════════════════════════════════════════
# CALIBRATION
# ═══════════════════════════════════════════════════════════════════════════════

def calibrate(frame_orig, scale, grid):
    """
    Isotropic calibration: click on two points with a KNOWN horizontal
    distance in the beam plane (typically the two supports = L_KNOWN_MM).

    Assumption: camera pixels are square  →  px_per_mm_x = px_per_mm_y.
    This is valid for any standard webcam facing the beam perpendicularly.

    The user enters the real distance in mm. A vertical reference can
    optionally be added afterwards via the menu if the camera is tilted.

    Returns: px_per_mm (single isotropic scale factor).
    """
    zv = ZoomView(frame_orig, scale)

    print("\n" + "═"*56)
    print(" CALIBRATION — two clicks on known horizontal distance")
    print("═"*56)
    print(f"  Click on the two beam endpoints (or any two points")
    print(f"  with a KNOWN horizontal distance in the beam plane).")
    print(f"  Suggested: fixed end and free end of beam = {L_KNOWN_MM:.0f} mm")
    print(f"  Then type the real distance when prompted.")

    pts = _two_clicks(frame_orig, scale, grid, zv,
                      "CALIBRATION: click points 1 and 2 on known distance",
                      color=(0, 255, 128))

    dx_px = abs(pts[1][0] - pts[0][0])
    dy_px = abs(pts[1][1] - pts[0][1])
    d_px  = np.hypot(dx_px, dy_px)

    # If the user clicked the beam endpoints, we know the distance = L_KNOWN_MM
    # but we allow them to enter any value for generality.
    while True:
        try:
            s = input(f"  Real distance between the two clicked points [mm] "
                      f"(ENTER = {L_KNOWN_MM:.0f} mm): ").strip()
            dist_mm = float(s) if s else L_KNOWN_MM
            if dist_mm <= 0: raise ValueError
            break
        except ValueError:
            print("  Please enter a positive number.")

    px_per_mm = d_px / dist_mm
    print(f"  -> {d_px:.1f} px  =  {dist_mm:.1f} mm")
    print(f"  -> px_per_mm = {px_per_mm:.4f}  (isotropic, both axes)")

    grid.calibrate(px_per_mm * scale)   # same scale for X and Y
    grid.set_origin(int(pts[0][0]*scale), int(pts[0][1]*scale))

    # Show result
    img_ok = zv.render(); grid.draw(img_ok)
    cv2.putText(img_ok,
        f"Calibration OK: {px_per_mm:.4f} px/mm  |  step={grid.step_mm} mm  |  SPACE: continue",
        (10,28), cv2.FONT_HERSHEY_SIMPLEX, 0.50, (0,255,128), 2)
    cv2.imshow(WIN, img_ok)
    while True:
        k = cv2.waitKey(30) & 0xFF
        if k == 27: cv2.destroyAllWindows(); sys.exit(0)
        if k == ord(' '): break
        if grid.handle_key(k):
            img_ok2 = zv.render(); grid.draw(img_ok2)
            cv2.putText(img_ok2,
                f"Calibration OK: {px_per_mm:.4f} px/mm  |  step={grid.step_mm} mm  |  SPACE: continue",
                (10,28), cv2.FONT_HERSHEY_SIMPLEX, 0.50, (0,255,128), 2)
            cv2.imshow(WIN, img_ok2)

    return px_per_mm


def calibrate_vertical(frame_orig, scale, grid, px_per_mm_x):
    """
    Optional anisotropic vertical calibration.
    Use only if the camera is significantly tilted in the vertical plane
    and the two-click horizontal calibration is not sufficient.

    Requires a physical vertical reference of known length in the image
    (e.g. a ruler taped vertically at the support).

    Returns px_per_mm_y.
    """
    zv = ZoomView(frame_orig, scale)

    print("\n" + "═"*56)
    print(" VERTICAL CALIBRATION (optional — for tilted cameras)")
    print("═"*56)
    print("  Click on the top and bottom of a vertical reference")
    print("  of known length (e.g. ruler, marked distance on support).")

    pts = _two_clicks(frame_orig, scale, grid, zv,
                      "VERTICAL CAL: click top and bottom of known vertical distance",
                      color=(0, 140, 255))

    dy_px = abs(pts[1][1] - pts[0][1])
    while True:
        try:
            dist_mm = float(input("  Vertical distance [mm]: "))
            if dist_mm > 0: break
        except ValueError: pass
        print("  Enter a positive number.")

    px_per_mm_y = dy_px / dist_mm
    print(f"  -> {dy_px:.1f} px  =  {dist_mm:.1f} mm")
    print(f"  -> px_per_mm_y = {px_per_mm_y:.4f}")
    print(f"  -> Y/X ratio   = {px_per_mm_y/px_per_mm_x:.4f}  "
          f"({'< 1 expected for downward-tilted camera' if px_per_mm_y < px_per_mm_x else 'check setup'})")

    grid.calibrate(px_per_mm_x * scale, px_per_mm_y * scale)
    return px_per_mm_y


# ═══════════════════════════════════════════════════════════════════════════════
# NEUTRAL AXIS ADJUSTMENT
# ═══════════════════════════════════════════════════════════════════════════════

def adjust_neutral_axis(frame_loaded, frame_unloaded, y_init_px, scale, grid):
    """
    Interactive adjustment of the neutral axis reference line.
    The semi-transparent overlay of the unloaded frame helps the user
    align the cyan line with the beam's horizontal reference position.

    Controls: W/UP — move up   S/DOWN — move down   Wheel — fine
              Q/E  — overlay opacity   G — grid   ENTER — confirm
    """
    H, W = frame_loaded.shape[:2]
    dw = int(W*scale); dh = int(H*scale)

    loaded_d   = cv2.resize(frame_loaded,   (dw, dh))
    unloaded_d = cv2.resize(frame_unloaded, (dw, dh))

    alpha = [0.55]
    yn_d  = [int(y_init_px * scale)]

    def redraw():
        blended = cv2.addWeighted(loaded_d, 1.0, unloaded_d, alpha[0], 0)
        if grid: grid.draw(blended)
        y = yn_d[0]
        cv2.line(blended, (0,y), (dw,y), (0,220,255), 2)
        for x in range(0, dw, 60):
            cv2.line(blended, (x, y-6), (x, y+6), (0,220,255), 2)
        cv2.putText(blended,
            "NEUTRAL AXIS: align cyan line with horizontal reference (no load, no gravity)",
            (10,26), cv2.FONT_HERSHEY_SIMPLEX, 0.48, (0,220,255), 2)
        cv2.putText(blended,
            "W/UP: up  S/DOWN: down  Wheel: fine  Q/E: opacity  G: grid  Enter: confirm",
            (10,46), cv2.FONT_HERSHEY_SIMPLEX, 0.37, (0,220,255), 1)
        cv2.putText(blended, f"Unloaded overlay: {int(alpha[0]*100)}%",
            (dw-230, 26), cv2.FONT_HERSHEY_SIMPLEX, 0.40, (200,200,0), 1)
        cv2.imshow(WIN, blended)

    def cb(event, x, y, flags, param):
        moved = False
        if grid and grid.handle_mouse_drag(event, x, y, flags):
            moved = True
        if event == cv2.EVENT_MOUSEWHEEL:
            yn_d[0] = int(np.clip(yn_d[0] - (1 if flags>0 else -1), 0, dh-1))
            moved = True
        if moved: redraw()

    cv2.setMouseCallback(WIN, cb)
    redraw()
    print("\n[NEUTRAL AXIS] Align the cyan line with the beam's reference position.")
    print("  The faint image is the unloaded beam.")
    print("  W/S or mouse wheel to move.  ENTER to confirm.")

    while True:
        k = cv2.waitKey(20) & 0xFF
        if k == 27:  cv2.destroyAllWindows(); sys.exit(0)
        if k in (13, 10): break
        if k in (ord('w'), ord('W'), 82): yn_d[0] = max(0, yn_d[0]-1); redraw()
        elif k in (ord('s'), ord('S'), 84):
            yn_d[0] = min(dh-1, yn_d[0]+1); redraw()
        elif k in (ord('q'), ord('Q')): alpha[0] = min(0.9, alpha[0]+0.05); redraw()
        elif k in (ord('e'), ord('E')): alpha[0] = max(0.0, alpha[0]-0.05); redraw()
        elif grid and grid.handle_key(k): redraw()

    cv2.setMouseCallback(WIN, lambda *a: None)
    y_adj = yn_d[0] / scale
    print(f"  -> Neutral axis at y = {y_adj:.1f} px (original frame)")
    return y_adj


# ═══════════════════════════════════════════════════════════════════════════════
# AUTO CONTOUR DETECTION
# ═══════════════════════════════════════════════════════════════════════════════

def auto_detect_beam(frame_orig, y_neutro_px, x_start_px, x_end_px):
    """
    Canny-based edge detection along the beam axis.
    Searches in a band of ±AUTO_BAND_HALF pixels around the neutral axis.
    Returns list of (x,y) tuples in original pixel coordinates.

    Note: this method is not robust for low-contrast materials (e.g. pine wood
    without colour markers). Manual digitisation is recommended in those cases.
    """
    H, W  = frame_orig.shape[:2]
    gray  = cv2.cvtColor(frame_orig, cv2.COLOR_BGR2GRAY)
    y0 = max(0, int(y_neutro_px) - AUTO_BAND_HALF)
    y1 = min(H, int(y_neutro_px) + AUTO_BAND_HALF)
    x0 = max(0, int(min(x_start_px, x_end_px)))
    x1 = min(W, int(max(x_start_px, x_end_px)))

    band  = gray[y0:y1, x0:x1].copy()
    edges = cv2.Canny(band, AUTO_CANNY_LOW, AUTO_CANNY_HIGH)
    edges = cv2.dilate(edges, np.ones((3,3), np.uint8), iterations=1)

    raw_xs, raw_ys = [], []
    ref_row = int(y_neutro_px) - y0
    for col in range(edges.shape[1]):
        rows = np.where(edges[:, col] > 0)[0]
        if not len(rows): continue
        above = rows[rows <= ref_row]
        row   = above[-1] if len(above) > 0 else rows[0]
        raw_xs.append(x0 + col)
        raw_ys.append(float(y0 + row))

    if len(raw_ys) < 10:
        print("[AUTO] Too few points detected."); return []

    ys     = np.array(raw_ys)
    k      = AUTO_MEDFILT_K
    if k > len(ys):
        k = max(3, len(ys) if len(ys)%2==1 else len(ys)-1)
    half_k = k // 2
    ys_filt = np.array([
        np.median(ys[max(0,i-half_k):min(len(ys),i+half_k+1)])
        for i in range(len(ys))
    ])
    step = max(1, len(raw_xs) // 60)
    pts  = [(int(raw_xs[i]), int(ys_filt[i]))
            for i in range(0, len(raw_xs), step)]
    print(f"[AUTO] {len(pts)} points detected.")
    return pts


# ═══════════════════════════════════════════════════════════════════════════════
# DIGITISATION
# ═══════════════════════════════════════════════════════════════════════════════

def digitize_deflection(frame_orig, frame_unloaded, y_neutro_adj,
                        scale, px_per_mm, x_fix_px, grid):
    """
    Interactive point digitisation on the loaded beam image.

    Controls:
      Left-click — add point
      A          — auto-detect contour (Canny)
      O          — toggle unloaded overlay
      Z          — undo last point
      G          — toggle grid
      ENTER      — accept and proceed
      ESC        — exit program
    """
    H, W  = frame_orig.shape[:2]
    dw = int(W*scale); dh = int(H*scale)

    img_d = cv2.resize(frame_orig,     (dw, dh))
    unl_d = cv2.resize(frame_unloaded, (dw, dh))
    base  = img_d.copy()

    yn = int(y_neutro_adj * scale)
    cv2.line(base, (0,yn), (dw,yn), (0,220,255), 2)
    for x in range(0, dw, 60):
        cv2.line(base, (x,yn-6), (x,yn+6), (0,220,255), 2)
    cv2.putText(base, "Neutral axis",
                (10, yn-8), cv2.FONT_HERSHEY_SIMPLEX, 0.44, (0,220,255), 1)
    cv2.putText(base,
        "Left-click: point  |  A: auto  |  O: overlay  |  G: grid  |  Z: undo  |  Enter: OK",
        (10,28), cv2.FONT_HERSHEY_SIMPLEX, 0.42, (0,255,255), 2)

    pts          = []
    show_overlay = [False]

    def redraw():
        canvas = (cv2.addWeighted(base, 1.0, unl_d, 0.25, 0)
                  if show_overlay[0] else base.copy())
        if grid: grid.draw(canvas)
        for xo, yo in pts:
            cv2.circle(canvas, (int(xo*scale), int(yo*scale)), 4, (0,255,0), -1)
        if len(pts) > 1:
            sp = sorted(pts, key=lambda p: p[0])
            for i in range(len(sp)-1):
                cv2.line(canvas,
                         (int(sp[i][0]*scale),   int(sp[i][1]*scale)),
                         (int(sp[i+1][0]*scale), int(sp[i+1][1]*scale)),
                         (0,200,0), 1)
        cv2.putText(canvas, f"Points: {len(pts)}",
                    (dw-160, 30), cv2.FONT_HERSHEY_SIMPLEX, 0.65, (255,255,0), 2)
        cv2.imshow(WIN, canvas)

    def cb(event, x, y, flags, param):
        moved = False
        if grid and grid.handle_mouse_drag(event, x, y, flags): moved = True
        if event == cv2.EVENT_LBUTTONDOWN:
            pts.append((x/scale, y/scale)); moved = True
        if moved: redraw()

    cv2.setMouseCallback(WIN, cb)
    redraw()
    print("\n  Left-click: manual point.  'A': auto-detect.  'O': overlay.")
    print("  Z: undo  |  Enter: compute  |  ESC: exit.")

    while True:
        k = cv2.waitKey(30) & 0xFF
        if k == 27: cv2.destroyAllWindows(); sys.exit(0)
        elif k in (ord('z'), ord('Z')):
            if pts: pts.pop(); redraw()
        elif k in (ord('o'), ord('O')):
            show_overlay[0] = not show_overlay[0]; redraw()
        elif k == ord('a'):
            print("  [AUTO] Detecting contour...")
            x_end = x_fix_px + int(L_KNOWN_MM * px_per_mm)
            ap = auto_detect_beam(frame_orig, y_neutro_adj, x_fix_px, x_end)
            if ap:
                pts.clear(); pts.extend(ap); redraw()
            else:
                print("  [AUTO] No result — try manual digitisation.")
        elif k in (13, 10):
            if len(pts) < 5:
                print("  At least 5 points required.")
            else:
                break
        elif grid and grid.handle_key(k): redraw()

    cv2.setMouseCallback(WIN, lambda *a: None)
    return list(pts)


# ─────────────────────────────────────────────────────────────────────────────
# HELPER: convert pixel points to mm coordinates
# ─────────────────────────────────────────────────────────────────────────────

def pts_to_mm(pts_px, x_fix_px, y_ref_adj, px_per_mm_x, px_per_mm_y, L_mm):
    """
    Convert a list of (x_px, y_px) original-frame pixel coordinates to
    (x_mm, v_mm) in beam coordinates.

    x_mm : distance from reference support along beam axis  [mm]
    v_mm : downward deflection from neutral axis reference  [mm]
            positive = downward (same sign convention as Euler-Bernoulli)

    Points are sorted by x_mm and clipped to [0, L_mm × 1.02].
    Returns (x_mm_arr, v_mm_arr) as numpy arrays.
    """
    arr      = np.array(pts_px)
    x_mm_all = np.abs(arr[:, 0] - x_fix_px) / px_per_mm_x
    v_mm_all = (arr[:, 1] - y_ref_adj)       / px_per_mm_y
    idx      = np.argsort(x_mm_all)
    x_mm_all = x_mm_all[idx]
    v_mm_all = v_mm_all[idx]
    mask     = x_mm_all <= L_mm * 1.02
    return x_mm_all[mask], v_mm_all[mask]


# ─────────────────────────────────────────────────────────────────────────────
# DIGITISE UNLOADED BEAM  (step 1 of each test)
# ─────────────────────────────────────────────────────────────────────────────

def digitize_unloaded(frame_unloaded, y_neutro_adj, scale,
                      px_per_mm, x_fix_px, grid):
    """
    Digitise the beam shape in the UNLOADED state (deflected by self-weight
    only). This forms the v=0 reference for the incremental measurement.

    Visually identical to digitize_deflection but with a different header
    colour (amber) and label so the user knows which step they are in.

    Returns list of (x_px, y_px) in original frame coordinates.
    """
    H, W  = frame_unloaded.shape[:2]
    dw = int(W*scale); dh = int(H*scale)

    img_d = cv2.resize(frame_unloaded, (dw, dh))
    base  = img_d.copy()

    yn = int(y_neutro_adj * scale)
    cv2.line(base, (0,yn), (dw,yn), (0,220,255), 1)
    cv2.putText(base,
        "STEP 1 — Digitise UNLOADED beam  |  Left-click: point  |  A: auto  |  Z: undo  |  Enter: OK",
        (10, 28), cv2.FONT_HERSHEY_SIMPLEX, 0.42, (0,200,255), 2)
    cv2.putText(base, "Click along the beam centreline (no load)",
        (10, 50), cv2.FONT_HERSHEY_SIMPLEX, 0.40, (200,200,0), 1)

    pts = []

    def redraw():
        canvas = base.copy()
        if grid: grid.draw(canvas)
        for xo, yo in pts:
            cv2.circle(canvas, (int(xo*scale), int(yo*scale)), 4, (0,200,255), -1)
        if len(pts) > 1:
            sp = sorted(pts, key=lambda p: p[0])
            for i in range(len(sp)-1):
                cv2.line(canvas,
                         (int(sp[i][0]*scale),   int(sp[i][1]*scale)),
                         (int(sp[i+1][0]*scale), int(sp[i+1][1]*scale)),
                         (0,160,200), 1)
        cv2.putText(canvas, f"Unloaded pts: {len(pts)}",
                    (dw-220, 30), cv2.FONT_HERSHEY_SIMPLEX, 0.65, (0,200,255), 2)
        cv2.imshow(WIN, canvas)

    def cb(event, x, y, flags, param):
        moved = False
        if grid and grid.handle_mouse_drag(event, x, y, flags): moved = True
        if event == cv2.EVENT_LBUTTONDOWN:
            pts.append((x/scale, y/scale)); moved = True
        if moved: redraw()

    cv2.setMouseCallback(WIN, cb)
    redraw()
    print("\n  STEP 1 — Digitise the UNLOADED beam (self-weight reference).")
    print("  Left-click: point  |  A: auto-detect  |  Z: undo  |  Enter: done")

    while True:
        k = cv2.waitKey(30) & 0xFF
        if k == 27: cv2.destroyAllWindows(); sys.exit(0)
        elif k in (ord('z'), ord('Z')):
            if pts: pts.pop(); redraw()
        elif k == ord('a'):
            print("  [AUTO] Detecting contour...")
            x_end = x_fix_px + int(L_KNOWN_MM * px_per_mm)
            ap = auto_detect_beam(frame_unloaded, y_neutro_adj, x_fix_px, x_end)
            if ap:
                pts.clear(); pts.extend(ap); redraw()
            else:
                print("  [AUTO] No result — use manual clicks.")
        elif k in (13, 10):
            if len(pts) < 5:
                print("  At least 5 points required.")
            else:
                break
        elif grid and grid.handle_key(k): redraw()

    cv2.setMouseCallback(WIN, lambda *a: None)
    return list(pts)


# ─────────────────────────────────────────────────────────────────────────────
# INTERPOLATE REFERENCE  (align unloaded curve onto loaded x grid)
# ─────────────────────────────────────────────────────────────────────────────

def interpolate_reference(x_unloaded, v_unloaded, x_loaded):
    """
    Linearly interpolate the unloaded deflection curve onto the x positions
    of the loaded digitisation so that the subtraction Δv = v_loaded − v_ref
    is point-wise.

    Points outside the unloaded x range are extrapolated with the nearest
    endpoint value (flat extrapolation) to avoid edge artefacts.
    """
    return np.interp(x_loaded, x_unloaded, v_unloaded,
                     left=v_unloaded[0], right=v_unloaded[-1])

# ═══════════════════════════════════════════════════════════════════════════════
# BEAM CASE SELECTOR
# ═══════════════════════════════════════════════════════════════════════════════

def _ask_mm_value(label, L_total, allow_zero=False):
    """Helper: ask a single mm value in an OpenCV window. Returns float."""
    buf = []

    def redraw():
        img2 = np.zeros((280, 900, 3), dtype=np.uint8); img2[:] = (25,25,25)
        cv2.putText(img2, label, (20,70),
                    cv2.FONT_HERSHEY_SIMPLEX, 0.55, (0,255,255), 2)
        cv2.putText(img2, f"Total beam length L = {L_total:.0f} mm",
                    (20,110), cv2.FONT_HERSHEY_SIMPLEX, 0.48, (150,150,150), 1)
        cv2.putText(img2, "".join(buf) + "_",
                    (20,185), cv2.FONT_HERSHEY_SIMPLEX, 1.00, (0,255,128), 2)
        cv2.putText(img2, "ENTER: confirm  |  ESC: exit",
                    (20,240), cv2.FONT_HERSHEY_SIMPLEX, 0.44, (150,150,150), 1)
        cv2.imshow(WIN, img2)

    redraw()
    while True:
        k = cv2.waitKey(50) & 0xFF
        if k == 27: cv2.destroyAllWindows(); sys.exit(0)
        if k in (13, 10):
            try:
                val = float("".join(buf))
                lo = 0.0 if allow_zero else 0.0
                if val > lo: return val
                buf.clear()
            except ValueError:
                pass
            redraw()
        elif k == 8 and buf: buf.pop(); redraw()
        elif chr(k) in '0123456789.': buf.append(chr(k)); redraw()


def select_mode_cv():
    """
    First screen shown at startup. User chooses:
      A — Live camera  (current behaviour, requires webcam)
      B — Static image (load a .jpg/.png file from disk)

    Returns 'live' or 'static'.
    """
    img = np.zeros((340, 780, 3), dtype=np.uint8)
    img[:] = (25, 25, 25)

    cv2.putText(img, "BEAM DEFLECTION & STRESS ANALYSIS",
                (20, 50), cv2.FONT_HERSHEY_SIMPLEX, 0.72, (0,255,255), 2)
    cv2.putText(img, "Select working mode:",
                (20, 100), cv2.FONT_HERSHEY_SIMPLEX, 0.55, (200,200,200), 1)

    cv2.putText(img, "A  —  Live camera  (real-time webcam capture)",
                (20, 155), cv2.FONT_HERSHEY_SIMPLEX, 0.58, (0,255,128), 1)
    cv2.putText(img, "B  —  Static image  (load .jpg / .png from disk)",
                (20, 210), cv2.FONT_HERSHEY_SIMPLEX, 0.58, (255,200,60), 1)

    cv2.putText(img, "Press  A  or  B   |   ESC: exit",
                (20, 290), cv2.FONT_HERSHEY_SIMPLEX, 0.50, (180,180,180), 1)
    cv2.imshow(WIN, img)

    while True:
        k = cv2.waitKey(50) & 0xFF
        if k == 27: cv2.destroyAllWindows(); sys.exit(0)
        if k in (ord('a'), ord('A')): return 'live'
        if k in (ord('b'), ord('B')): return 'static'


def load_static_image_cv():
    """
    Ask the user to type the full path to an image file in an OpenCV window.
    Returns the loaded frame (numpy array BGR) or exits on error/ESC.
    Falls back to tkinter file dialog if available.
    """
    # Try tkinter dialog first (much more user-friendly)
    try:
        import tkinter as tk
        from tkinter import filedialog
        root = tk.Tk(); root.withdraw()
        path = filedialog.askopenfilename(
            title="Select beam image",
            filetypes=[("Image files", "*.jpg *.jpeg *.png *.bmp *.tiff"), ("All files", "*.*")]
        )
        root.destroy()
        if not path:
            print("  [INFO] No file selected — returning to mode selection.")
            return None
    except Exception:
        # Fallback: type path manually in OpenCV window
        buf = []

        def redraw():
            img2 = np.zeros((240, 900, 3), dtype=np.uint8); img2[:] = (25,25,25)
            cv2.putText(img2, "Type full path to image file and press ENTER:",
                        (20,60), cv2.FONT_HERSHEY_SIMPLEX, 0.52, (0,255,255), 2)
            cv2.putText(img2, "".join(buf) + "_",
                        (20,140), cv2.FONT_HERSHEY_SIMPLEX, 0.65, (0,255,128), 2)
            cv2.putText(img2, "ESC: cancel",
                        (20,200), cv2.FONT_HERSHEY_SIMPLEX, 0.44, (150,150,150), 1)
            cv2.imshow(WIN, img2)

        redraw()
        while True:
            k = cv2.waitKey(50) & 0xFF
            if k == 27: return None
            if k in (13, 10):
                path = "".join(buf).strip(); break
            elif k == 8 and buf: buf.pop(); redraw()
            elif 32 <= k <= 126: buf.append(chr(k)); redraw()

    frame = cv2.imread(path)
    if frame is None:
        print(f"  [ERROR] Cannot read image: {path}")
        return None
    print(f"  [INFO] Loaded image: {path}  ({frame.shape[1]}×{frame.shape[0]} px)")
    return frame


# ─────────────────────────────────────────────────────────────────────────────
def select_beam_case_cv():
    """
    OpenCV menu to select beam case and geometric parameters.
    Returns (BeamCase, params_dict) where params_dict contains the
    relevant lengths for the chosen case.
    """
    img = np.zeros((620, 980, 3), dtype=np.uint8)
    img[:] = (25, 25, 25)

    cv2.putText(img, "BEAM DEFLECTION ANALYSIS 2 — Select beam case",
                (20, 46), cv2.FONT_HERSHEY_SIMPLEX, 0.68, (0,255,255), 2)

    entries = [
        ("1 — Cantilever, load at FREE END",
         (0,255,128)),
        ("2 — Cantilever, load at INTERMEDIATE point  (x = a)",
         (255,200,60)),
        ("3 — Simply supported, load at x = a",
         (100,180,255)),
        ("4 — Overhanging:  pin-A  roller-B  load-at-C (free end)",
         (255,140,80)),
        ("5 — Overhanging:  pin-A  load-at-B  roller-C  free-end-D",
         (180,120,255)),
        ("6 — Overhanging:  pin-A  roller-B  load-at-C  free-end-D",
         (80,220,180)),
    ]
    for i, (txt, col) in enumerate(entries):
        cv2.putText(img, txt, (20, 105+i*72),
                    cv2.FONT_HERSHEY_SIMPLEX, 0.54, col, 1)

    cv2.putText(img, "Press  1 / 2 / 3 / 4 / 5 / 6  to choose   |   ESC: exit",
                (20, 555), cv2.FONT_HERSHEY_SIMPLEX, 0.52, (180,180,180), 1)
    cv2.imshow(WIN, img)

    case_map = {
        ord('1'): BeamCase.CANTILEVER_FREE_END,
        ord('2'): BeamCase.CANTILEVER_INTERMEDIATE,
        ord('3'): BeamCase.SIMPLY_SUPPORTED,
        ord('4'): BeamCase.OVERHANG_RIGHT_FREE_END,
        ord('5'): BeamCase.OVERHANG_LOAD_BETWEEN,
        ord('6'): BeamCase.OVERHANG_LOAD_IN_SPAN,
    }
    case = None
    while case is None:
        k = cv2.waitKey(50) & 0xFF
        if k == 27: cv2.destroyAllWindows(); sys.exit(0)
        case = case_map.get(k)

    L = L_KNOWN_MM
    params = {}

    if case == BeamCase.CANTILEVER_FREE_END:
        params = {'a_mm': L}

    elif case == BeamCase.CANTILEVER_INTERMEDIATE:
        a = _ask_mm_value("Load position  a  from FIXED END [mm]:", L)
        params = {'a_mm': a}

    elif case == BeamCase.SIMPLY_SUPPORTED:
        a = _ask_mm_value("Load position  a  from LEFT SUPPORT [mm]:", L)
        params = {'a_mm': a}

    elif case == BeamCase.OVERHANG_RIGHT_FREE_END:
        # Beam: A---B---C   total = La+Lb = L_KNOWN_MM
        La = _ask_mm_value("Distance  La  from pin A to roller B [mm]:", L)
        Lb = L - La
        print(f"  -> Lb (B to free end C) = {Lb:.1f} mm")
        params = {'La': La, 'Lb': Lb, 'a_mm': L}

    elif case == BeamCase.OVERHANG_LOAD_BETWEEN:
        # Beam: A---B---C---D   La+Lb+Lc = L_KNOWN_MM
        La = _ask_mm_value("Distance  La  from pin A to load point B [mm]:", L)
        Lb = _ask_mm_value("Distance  Lb  from load B to roller C [mm]:", L - La)
        Lc = L - La - Lb
        print(f"  -> Lc (C to free end D) = {Lc:.1f} mm")
        params = {'La': La, 'Lb': Lb, 'Lc': Lc, 'a_mm': La}

    elif case == BeamCase.OVERHANG_LOAD_IN_SPAN:
        # Beam: A---B---C---D   La+Lb+Lc = L_KNOWN_MM
        La = _ask_mm_value("Distance  La  from pin A to roller B [mm]:", L)
        Lb = _ask_mm_value("Distance  Lb  from roller B to load C [mm]:", L - La)
        Lc = L - La - Lb
        print(f"  -> Lc (C to free end D) = {Lc:.1f} mm")
        params = {'La': La, 'Lb': Lb, 'Lc': Lc, 'a_mm': La+Lb}

    labels = {
        BeamCase.CANTILEVER_FREE_END:     "Cantilever — load at free end",
        BeamCase.CANTILEVER_INTERMEDIATE: "Cantilever — load at intermediate point",
        BeamCase.SIMPLY_SUPPORTED:        "Simply supported",
        BeamCase.OVERHANG_RIGHT_FREE_END: "Overhanging — load at right free end",
        BeamCase.OVERHANG_LOAD_BETWEEN:   "Overhanging — load between supports",
        BeamCase.OVERHANG_LOAD_IN_SPAN:   "Overhanging — load in right overhang",
    }
    print(f"\n  Beam case: {labels[case]}")
    print(f"  Parameters: {params}")
    return case, params



def ask_mass_cv(test_id):
    """Keyboard input for real mass [g] in OpenCV window. Returns float or None."""
    buf = []

    def redraw():
        img = np.zeros((220, 700, 3), dtype=np.uint8); img[:] = (30,30,30)
        cv2.putText(img, f"TEST {test_id} — Real mass of the load (grams)",
                    (20,50), cv2.FONT_HERSHEY_SIMPLEX, 0.58, (0,255,255), 2)
        cv2.putText(img, "Type grams and press ENTER  |  ESC = skip",
                    (20,85), cv2.FONT_HERSHEY_SIMPLEX, 0.44, (150,150,150), 1)
        cv2.putText(img, "".join(buf) + "_",
                    (20,155), cv2.FONT_HERSHEY_SIMPLEX, 1.00, (0,255,128), 2)
        cv2.imshow(WIN, img)

    redraw()
    while True:
        k = cv2.waitKey(50) & 0xFF
        if k == 27: return None
        if k in (13, 10):
            try:    return float("".join(buf)) if buf else None
            except: buf.clear(); redraw()
        if k == 8 and buf: buf.pop(); redraw()
        if chr(k) in '0123456789.': buf.append(chr(k)); redraw()


def show_menu_cv(test_id, results, case_label):
    """Main session menu. Returns 'test', 'quit', or 'recal'."""
    img = np.zeros((460, 1000, 3), dtype=np.uint8); img[:] = (25,25,25)
    cv2.putText(img, "BEAM DEFLECTION ANALYSIS  —  Session in progress",
                (20, 45), cv2.FONT_HERSHEY_SIMPLEX, 0.70, (0,255,255), 2)
    cv2.putText(img, f"Case: {case_label}",
                (20, 80), cv2.FONT_HERSHEY_SIMPLEX, 0.52, (180,180,180), 1)
    cv2.putText(img, f"Tests completed: {len(results)}",
                (20, 110), cv2.FONT_HERSHEY_SIMPLEX, 0.52, (180,180,180), 1)

    controls = [
        (f"ENTER / N   ->  New test  (TEST {test_id})",    (0,255,128)),
        ( "R           ->  Recalibrate (new unloaded frame)", (255,200,0)),
        ( "Q / ESC     ->  End session and print summary",    (200,100,100)),
    ]
    for i, (txt, col) in enumerate(controls):
        cv2.putText(img, txt, (20, 160+i*40),
                    cv2.FONT_HERSHEY_SIMPLEX, 0.56, col, 1)

    if results:
        r = results[-1]
        cv2.putText(img, f"Last result  (TEST {r['test_id']}):",
                    (20, 310), cv2.FONT_HERSHEY_SIMPLEX, 0.50, (200,200,200), 1)
        cv2.putText(img,
            f"  P = {r['P_N']:.4f} N  (~{r['mass_g_est']:.1f} g)   "
            f"delta_load = {r['delta_load']:.1f} mm   delta_total = {r['delta_total']:.1f} mm",
            (20, 342), cv2.FONT_HERSHEY_SIMPLEX, 0.50, (200,200,200), 1)
        if r['mass_g_real']:
            P_real = r['mass_g_real'] * 9.81 / 1000
            err    = (r['P_N'] - P_real) / P_real * 100
            cv2.putText(img,
                f"  Real mass: {r['mass_g_real']:.0f} g   Error P: {err:+.1f}%",
                (20, 372), cv2.FONT_HERSHEY_SIMPLEX, 0.50, (200,200,200), 1)

    cv2.imshow(WIN, img)
    while True:
        k = cv2.waitKey(50) & 0xFF
        if k in (13, ord('n'), ord('N')): return 'test'
        if k in (27, ord('q'), ord('Q')): return 'quit'
        if k in (ord('r'), ord('R')):     return 'recal'


# ═══════════════════════════════════════════════════════════════════════════════
# PLOT RESULTS
# ═══════════════════════════════════════════════════════════════════════════════

def plot_results(x_mm, dv_mm, v_loaded_mm, v_unloaded_mm,
                 v_func, M_func, V_func,
                 L_mm, P_N, a_mm, case,
                 img_overlay=None, save_path=None):
    """
    Single-column A4-portrait report figure.

    DISPLAY / AESTHETICS ONLY — the analysis is untouched. All computed
    quantities (x_crit, M_crit, sigma, fitted curves, residuals) are identical
    to the previous version; only the figure layout, fonts and labels changed.

    Layout (top to bottom):
      [camera image]                                  (if available)
      Deflection v(x)            (full width)
      Bending Moment M(x)        (full width, arrow at M_max)
      Shear Force V(x)           (full width)
      Normal Stress @ x_crit (left)  |  Key results box (right)
      Fit Residuals              (full width)
    """
    # ── Figure size tuned to A4 portrait usable area (~2.5 cm margins). ───────
    #    Adjust only these two numbers if your template margins differ.
    FIG_W, FIG_H = 8.0, 11.5

    # ── Computations (UNCHANGED) ──────────────────────────────────────────────
    x_plot = np.linspace(0, L_mm, 600)
    v_p    = v_func(x_plot)
    M_p    = M_func(x_plot)
    V_p    = V_func(x_plot)

    idx_crit  = int(np.argmax(np.abs(M_p)))
    x_crit    = float(x_plot[idx_crit])
    M_crit    = float(M_p[idx_crit])
    sigma_max = abs(M_crit) * (h_mm / 2.0) / I_mm4   # [MPa]

    resid = dv_mm - v_func(x_mm)
    rms   = float(np.sqrt(np.mean(resid**2)))
    V_max = float(np.max(np.abs(V_p)))

    has_img = img_overlay is not None
    xmax    = max(L_mm, float(np.max(x_mm))) if len(x_mm) else L_mm

    # ── Figure + grid (single column; stress row split 56 / 44) ───────────────
    fig = plt.figure(figsize=(FIG_W, FIG_H), facecolor='white')

    if has_img:
        h_ratios = [2.8, 1.45, 1.15, 1.00, 1.30, 0.72]
        nrows    = 6
    else:
        h_ratios = [1.45, 1.15, 1.00, 1.30, 0.72]
        nrows    = 5

    gs = fig.add_gridspec(
        nrows, 2,
        height_ratios=h_ratios,
        width_ratios=[0.56, 0.44],
        hspace=1.0, wspace=0.18,
        left=0.115, right=0.95, top=0.815, bottom=0.04
    )

    case_labels = {
        BeamCase.CANTILEVER_FREE_END:     "Cantilever — load at free end",
        BeamCase.CANTILEVER_INTERMEDIATE: f"Cantilever — load at a={a_mm:.0f} mm",
        BeamCase.SIMPLY_SUPPORTED:        f"Simply supported — load at a={a_mm:.0f} mm",
        BeamCase.OVERHANG_RIGHT_FREE_END: f"Overhanging — load at right free end",
        BeamCase.OVERHANG_LOAD_BETWEEN:   f"Overhanging — load between supports",
        BeamCase.OVERHANG_LOAD_IN_SPAN:   f"Overhanging — load in right overhang",
    }

    # ── Big bold title + non-bold parameter box (with margin) ─────────────────
    fig.suptitle(f"Deflection Analysis — {case_labels[case]}",
                 x=0.53, fontsize=15, fontweight='bold', y=0.987)

    P_sci = P_N / 1e-2   # express P in units of 1e-2 N (scientific notation)
    param_str = (
        f"b = {b_mm:.0f} mm     h = {h_mm:.0f} mm     "
        f"E = {E_MPa:.0f} MPa     EI = {EI:.0f} N·mm²\n"
        rf"L = {L_mm:.1f} mm          "
        rf"P = {P_sci:.2f}$\times 10^{{-2}}$ N   (~{P_N/9.81*1000:.1f} g)"
    )
    fig.text(0.53, 0.945, param_str, ha='center', va='top',
             fontsize=12, linespacing=1.7,
             bbox=dict(boxstyle='round,pad=0.7', facecolor='#f7f7f7',
                       edgecolor='#aaaaaa', linewidth=1.0))

    # ── Shared panel styling ──────────────────────────────────────────────────
    def style(ax, title, xl, yl):
        ax.set_title(title, fontsize=12, fontweight='bold', pad=8)
        ax.set_xlabel(xl, fontsize=12.5, labelpad=4)
        ax.set_ylabel(yl, fontsize=12.5, labelpad=4)
        ax.tick_params(labelsize=11)
        ax.grid(True, ls='--', alpha=0.4)

    r = 0
    # ── Camera image ──────────────────────────────────────────────────────────
    if has_img:
        ax_i = fig.add_subplot(gs[0, :])
        ax_i.imshow(cv2.cvtColor(img_overlay, cv2.COLOR_BGR2RGB))
        ax_i.set_title("Camera image with digitized points and analytical fit",
                       fontsize=11, pad=16)
        ax_i.axis('off')
        r = 1

    # ── Deflection v(x) ───────────────────────────────────────────────────────
    ax_v = fig.add_subplot(gs[r, :]); r += 1
    v_unl_interp_on_plot = np.interp(x_plot, x_mm, v_unloaded_mm)
    v_total_fit_mm       = v_unl_interp_on_plot + v_p
    ax_v.plot(x_mm, v_loaded_mm,   '^', color='#00B000', ms=5, alpha=0.7,
              label='v_loaded (digitized)', zorder=3)
    ax_v.plot(x_mm, v_unloaded_mm, 's', color='#E0A000', ms=5, alpha=0.7,
              label='v_unloaded (ref.)', zorder=3)
    ax_v.plot(x_plot, v_total_fit_mm, '-', color='#FF8C00', lw=2.4,
              label='Euler–Bernoulli fit')
    if a_mm is not None and 0 < a_mm < L_mm:
        ax_v.axvline(a_mm, color='magenta', ls=':', lw=1.6,
                     label=f'x = a = {a_mm:.0f} mm')
    ax_v.axvline(L_mm, color='red', ls='--', lw=1.6, label=f'x = L = {L_mm:.0f} mm')
    style(ax_v, "Deflection  v(x)", "x [mm]", "v(x) [mm]")
    ax_v.set_title("Deflection  v(x)", fontsize=12, fontweight='bold', pad=32)
    ax_v.set_xlabel(""); ax_v.tick_params(labelbottom=False)
    # small padding so the curve fills the panel; legend goes OUTSIDE (above the
    # panel) -> robust even for overhang shapes that rise and fall within the span
    _allv = np.concatenate([np.asarray(v_loaded_mm, float),
                            np.asarray(v_unloaded_mm, float), v_total_fit_mm])
    _vlo, _vhi = float(np.nanmin(_allv)), float(np.nanmax(_allv))
    _pad = 0.12 * max(_vhi - _vlo, 1e-6)
    ax_v.set_ylim(_vhi + _pad, _vlo - _pad)   # inverted: high v at bottom
    ax_v.legend(loc='lower center', bbox_to_anchor=(0.5, 1.0), ncol=3,
                fontsize=8, frameon=False, handletextpad=0.4,
                columnspacing=1.4, borderaxespad=0.2)

    # ── Bending Moment M(x) + arrow to M_max ──────────────────────────────────
    ax_M = fig.add_subplot(gs[r, :], sharex=ax_v); r += 1
    ax_M.plot(x_plot, M_p, 'r-', lw=2.6)
    ax_M.fill_between(x_plot, M_p, alpha=0.12, color='red')
    ax_M.axhline(0, color='0.6', lw=0.8, zorder=0)
    ax_M.axvline(x_crit, color='darkred', ls=':', lw=1.2)
    _mlo, _mhi = min(0.0, float(M_p.min())), max(0.0, float(M_p.max()))
    _mpad = 0.14 * max(_mhi - _mlo, 1e-6)
    ax_M.set_ylim(_mhi + _mpad, _mlo - _mpad)   # inverted + margin (0 stays off the frame)
    style(ax_M, "Bending Moment  M(x)", "x [mm]", "M(x) [N·mm]")
    # label in the empty (M~0) corner opposite the peak; thin arrow to the peak
    if x_crit <= L_mm / 2.0:
        tx, ha = 0.985, 'right'
    else:
        tx, ha = 0.015, 'left'
    ax_M.annotate(f"M_max = {M_crit:.1f} N·mm",
                  xy=(x_crit, M_crit), xycoords='data',
                  xytext=(tx, 0.10), textcoords='axes fraction',
                  ha=ha, va='bottom', fontsize=11.5, color='darkred',
                  arrowprops=dict(arrowstyle='->', color='darkred', lw=1.4,
                                  shrinkA=3, shrinkB=3))
    ax_M.set_xlabel(""); ax_M.tick_params(labelbottom=False)

    # ── Shear Force V(x) ──────────────────────────────────────────────────────
    ax_V = fig.add_subplot(gs[r, :], sharex=ax_v); r += 1
    ax_V.plot(x_plot, V_p, color='darkorange', lw=2.6)
    ax_V.fill_between(x_plot, V_p, alpha=0.12, color='darkorange')
    ax_V.axhline(0, color='0.6', lw=0.8, zorder=0)
    _vlo2, _vhi2 = min(0.0, float(V_p.min())), max(0.0, float(V_p.max()))
    _vpad = 0.18 * max(_vhi2 - _vlo2, 1e-6)
    ax_V.set_ylim(_vlo2 - _vpad, _vhi2 + _vpad)   # margin so 0 is off the frame
    style(ax_V, "Shear Force  V(x)", "x [mm]", "V(x) [N]")

    # keep the three x-based diagrams aligned on a common x-axis
    ax_v.set_xlim(0, xmax)

    # ── Normal Stress cross-section (left of the stress row) ──────────────────
    ax_s = fig.add_subplot(gs[r, 0])
    y_vals  = np.linspace(-h_mm/2, h_mm/2, 300)
    sigma_y = -M_crit * y_vals / I_mm4   # signed: + tension, - compression
    ax_s.fill_betweenx(y_vals, 0, sigma_y, where=(sigma_y >= 0),
                       color='#2980b9', alpha=0.35, label='Tension (+)')
    ax_s.fill_betweenx(y_vals, 0, sigma_y, where=(sigma_y < 0),
                       color='#c0392b', alpha=0.35, label='Compression (−)')
    ax_s.plot(sigma_y, y_vals, 'k-', lw=2.3)
    ax_s.axhline(0, color='k', lw=1.2, ls='--', label='Neutral axis')
    ax_s.axvline(0, color='k', lw=0.8)
    rx0 = min(sigma_y.min(), 0) * 1.05
    rx1 = max(sigma_y.max(), 0) * 1.05
    ax_s.add_patch(plt.Rectangle((rx0, -h_mm/2), rx1 - rx0, h_mm,
                                 fill=False, edgecolor='#555', lw=1.2))
    ax_s.set_xlabel("σ [MPa]", fontsize=12.5, labelpad=4)
    ax_s.set_ylabel("y [mm]", fontsize=12, labelpad=4)
    ax_s.set_title(f"Normal Stress @ x = {x_crit:.0f} mm",
                   fontsize=12, fontweight='bold', pad=8)
    ax_s.tick_params(labelsize=11)
    ax_s.grid(True, ls='--', alpha=0.4)
    ax_s.legend(fontsize=9, framealpha=0.92, handletextpad=0.4,
                loc=('upper left' if M_crit >= 0 else 'upper right'))

    # ── Key results box (right of the stress row) ─────────────────────────────
    ax_b = fig.add_subplot(gs[r, 1]); r += 1
    ax_b.axis('off')
    res_lines = [
        "Key results",
        f"M_crit = {M_crit:.2f} N·mm  (x = {x_crit:.0f} mm)",
        f"|V|_max = {V_max:.4f} N",
        f"σ_max = ±{sigma_max:.3f} MPa",
        f"RMS = {rms:.3f} mm",
    ]
    ax_b.text(0.02, 0.95, "\n".join(res_lines),
              transform=ax_b.transAxes, va='top', ha='left',
              fontsize=10.5, linespacing=1.55,
              bbox=dict(boxstyle='round,pad=0.6', facecolor='lightyellow',
                        edgecolor='#bbbbbb', linewidth=1.0))

    # ── Fit Residuals ─────────────────────────────────────────────────────────
    ax_r = fig.add_subplot(gs[r, :])
    bw   = max(L_mm / max(len(x_mm), 1) * 0.7, 1)
    ax_r.bar(x_mm, resid, width=bw, color='mediumpurple', alpha=0.78)
    ax_r.axhline(0, color='k', lw=0.8)
    ax_r.set_xlim(0, xmax)
    style(ax_r, f"Fit Residuals  RMS = {rms:.3f} mm", "x [mm]", "residual [mm]")

    if save_path:
        plt.savefig(save_path, dpi=200, facecolor='white')
        print(f"[INFO] Figure saved: {save_path}")
    plt.show()


# ═══════════════════════════════════════════════════════════════════════════════
# SINGLE TEST
# ═══════════════════════════════════════════════════════════════════════════════

def run_test(cap, scale, grid, frame_unloaded,
             px_per_mm_x, px_per_mm_y,
             support_px, y_ref_adj,
             test_id, case, params):
    """
    Execute one measurement test using the TWO-DIGITISATION strategy.
    'params' is the dict returned by select_beam_case_cv().
    """
    L_mm     = float(L_KNOWN_MM)
    x_fix_px = support_px[0]
    a_mm     = params.get('a_mm', L_mm)   # load position (for reference point)

    # ── Ask for real mass ─────────────────────────────────────────────────────
    mass_g = ask_mass_cv(test_id)
    print(f"  [TEST {test_id}] Real mass: "
          + (f"{mass_g:.1f} g" if mass_g else "not entered"))

    # ── Print self-weight info (informational only) ───────────────────────────
    v_sw_tip = v_self_weight_analytical(np.array([L_mm]), L_mm)[0]
    print(f"  [INFO] Analytical self-weight tip deflection: {v_sw_tip:.1f} mm "
          f"(subtracted empirically, not analytically)")

    # ── Instructions ──────────────────────────────────────────────────────────
    img_w = np.zeros((360, 1020, 3), dtype=np.uint8); img_w[:] = (25,25,25)
    cv2.putText(img_w, f"TEST {test_id}  —  Two-step digitisation",
                (20, 55), cv2.FONT_HERSHEY_SIMPLEX, 0.70, (0,255,255), 2)
    for i, (txt, col) in enumerate([
        ("STEP 1: Digitise the UNLOADED beam (no weight on it yet)", (0,200,255)),
        ("STEP 2: Place the load, wait for equilibrium",              (180,180,180)),
        ("STEP 3: Capture frame and digitise the LOADED beam",        (0,255,128)),
        ("  Press  SPACE  to start STEP 1",                           (0,255,128)),
        ("  ESC = cancel this test",                                  (200,100,100)),
    ]):
        cv2.putText(img_w, txt, (20, 110+i*48),
                    cv2.FONT_HERSHEY_SIMPLEX, 0.52, col, 1)
    cv2.imshow(WIN, img_w)
    while True:
        k = cv2.waitKey(50) & 0xFF
        if k == 27: print("  Test cancelled."); return None
        if k == ord(' '): break

    # ══════════════════════════════════════════════════════════════════════════
    # STEP 1 — Digitise unloaded beam
    # ══════════════════════════════════════════════════════════════════════════
    print(f"\n══ TEST {test_id} — STEP 1: DIGITISE UNLOADED BEAM ══════")
    pts_unloaded_px = digitize_unloaded(
        frame_unloaded, y_ref_adj, scale,
        px_per_mm_x, x_fix_px, grid)

    x_unl, v_unl = pts_to_mm(pts_unloaded_px, x_fix_px, y_ref_adj,
                               px_per_mm_x, px_per_mm_y, L_mm)

    if len(x_unl) < 5:
        print("[ERROR] Too few unloaded points."); return None

    print(f"  Unloaded digitised: {len(x_unl)} points  "
          f"| tip v_unloaded = {v_unl[-1]:.1f} mm  "
          f"(analytical: {v_sw_tip:.1f} mm)")

    # ══════════════════════════════════════════════════════════════════════════
    # STEP 2 — Capture loaded frame
    # ══════════════════════════════════════════════════════════════════════════
    img_w2 = np.zeros((260, 900, 3), dtype=np.uint8); img_w2[:] = (25,25,25)
    cv2.putText(img_w2, "STEP 2 — Place the load on the beam",
                (20, 70), cv2.FONT_HERSHEY_SIMPLEX, 0.70, (0,255,255), 2)
    cv2.putText(img_w2, "Wait for static equilibrium, then press SPACE to capture",
                (20,120), cv2.FONT_HERSHEY_SIMPLEX, 0.55, (0,255,128), 1)
    cv2.putText(img_w2, "ESC = cancel",
                (20,165), cv2.FONT_HERSHEY_SIMPLEX, 0.50, (200,100,100), 1)
    cv2.imshow(WIN, img_w2)
    while True:
        k = cv2.waitKey(50) & 0xFF
        if k == 27: print("  Test cancelled."); return None
        if k == ord(' '): break

    frame_loaded = freeze_frame(cap, scale,
                                f"TEST {test_id}: WITH LOAD -> SPACE", grid)

    # ══════════════════════════════════════════════════════════════════════════
    # STEP 3 — Digitise loaded beam
    # ══════════════════════════════════════════════════════════════════════════
    print(f"\n══ TEST {test_id} — STEP 3: DIGITISE LOADED BEAM ══════════")

    # Update digitize header to say STEP 3
    pts_loaded_px = digitize_deflection(
        frame_loaded, frame_unloaded, y_ref_adj, scale,
        px_per_mm_x, x_fix_px, grid)

    x_load, v_load = pts_to_mm(pts_loaded_px, x_fix_px, y_ref_adj,
                                px_per_mm_x, px_per_mm_y, L_mm)

    if len(x_load) < 5:
        print("[ERROR] Too few loaded points."); return None

    L_dig = float(x_load[-1])
    if abs(L_dig - L_mm) / L_mm * 100 > 5.0:
        print(f"  [WARNING] Digitised length = {L_dig:.1f} mm  "
              f"(expected {L_mm:.0f} mm). Check calibration or re-digitise.")

    # ══════════════════════════════════════════════════════════════════════════
    # STEP 4 — Empirical subtraction: Δv = v_loaded − v_unloaded
    # ══════════════════════════════════════════════════════════════════════════
    v_ref_interp = interpolate_reference(x_unl, v_unl, x_load)
    dv           = v_load - v_ref_interp    # incremental deflection due to load

    print(f"  Loaded tip v          = {v_load[-1]:.2f} mm")
    print(f"  Unloaded tip v (ref)  = {v_ref_interp[-1]:.2f} mm")
    print(f"  Incremental Δv tip    = {dv[-1]:.2f} mm")

    if np.max(np.abs(dv)) < 0.5:
        print("  [WARNING] Incremental deflection < 0.5 mm — "
              "check that the load was actually applied.")

    # ══════════════════════════════════════════════════════════════════════════
    # STEP 5 — Fit Euler-Bernoulli to Δv(x)
    # ══════════════════════════════════════════════════════════════════════════
    La = params.get('La', L_mm)
    Lb = params.get('Lb', 0.0)
    Lc = params.get('Lc', 0.0)

    if case == BeamCase.CANTILEVER_FREE_END:
        P_N, v_func, M_func, V_func = fit_cantilever_free_end(x_load, dv, L_mm)
    elif case == BeamCase.CANTILEVER_INTERMEDIATE:
        P_N, v_func, M_func, V_func = fit_cantilever_intermediate(x_load, dv, L_mm, a_mm)
    elif case == BeamCase.SIMPLY_SUPPORTED:
        P_N, v_func, M_func, V_func = fit_simply_supported(x_load, dv, L_mm, a_mm)
    elif case == BeamCase.OVERHANG_RIGHT_FREE_END:
        P_N, v_func, M_func, V_func = fit_overhang_right_free_end(x_load, dv, La, Lb)
    elif case == BeamCase.OVERHANG_LOAD_BETWEEN:
        P_N, v_func, M_func, V_func = fit_overhang_load_between(x_load, dv, La, Lb, Lc)
    elif case == BeamCase.OVERHANG_LOAD_IN_SPAN:
        P_N, v_func, M_func, V_func = fit_overhang_load_in_span(x_load, dv, La, Lb, Lc)

    if P_N is None: return None
    if P_N < 0 and case in (BeamCase.CANTILEVER_FREE_END,
                             BeamCase.CANTILEVER_INTERMEDIATE,
                             BeamCase.SIMPLY_SUPPORTED):
        print(f"  [WARNING] Negative P = {P_N:.4f} N — check digitisation order "
              "(loaded beam should be BELOW unloaded beam).")

    mass_est   = abs(P_N) / 9.81 * 1000
    x_ref_pt   = np.array([a_mm])
    v_load_max = float(v_func(x_ref_pt)[0])

    P_for_theory = (mass_g * 9.81 / 1000) if mass_g else P_N

    if case == BeamCase.CANTILEVER_FREE_END:
        v_theory = P_for_theory * L_mm**3 / (3*EI)
    elif case == BeamCase.CANTILEVER_INTERMEDIATE:
        v_theory = P_for_theory * a_mm**3 / (3*EI)
    elif case == BeamCase.SIMPLY_SUPPORTED:
        b_ss     = L_mm - a_mm
        v_theory = P_for_theory * a_mm**2 * b_ss**2 / (3*EI*L_mm)
    elif case == BeamCase.OVERHANG_RIGHT_FREE_END:
        # v at free end C = -PLb²·L/(3EI)  (downward, stored as positive)
        v_theory = P_for_theory * Lb**2 * L_mm / (3*EI)
    elif case == BeamCase.OVERHANG_LOAD_BETWEEN:
        # v at load point B:  va(La) = Lb·La·(La²-La²-2LaLb)/(6EI·Sab) ... use v_func
        v_theory = abs(float(
            P_for_theory * Lb * La * (La**2 + 0 - 2*La*Lb) / (6*EI*(La+Lb))
        )) if False else abs(v_load_max)   # use measured; theory stored for reference only
        v_theory = abs(float(
            -P_for_theory * La**2 * Lb**2 / (3*EI*(La+Lb))
        ))
    elif case == BeamCase.OVERHANG_LOAD_IN_SPAN:
        # v at load point C = -PLb²(La+Lb)/(3EI)
        v_theory = P_for_theory * Lb**2 * (La+Lb) / (3*EI)

    print(f"\n{'='*54}")
    print(f"  TEST {test_id}  RESULTS")
    print(f"  P (estimated)      = {P_N:.4f} N  (~{mass_est:.1f} g)")
    if mass_g:
        P_real = mass_g * 9.81 / 1000
        err_p  = (P_N - P_real) / P_real * 100
        err_v  = (abs(v_load_max) - v_theory) / v_theory * 100
        print(f"  P (real)           = {P_real:.4f} N  ({mass_g:.1f} g)  "
              f"|  Error P = {err_p:+.1f}%")
        print(f"  Δv measured        = {v_load_max:.2f} mm")
        print(f"  Δv theory (P_real) = {v_theory:.2f} mm  "
              f"|  Error Δv = {err_v:+.1f}%")
    else:
        print(f"  Δv at ref. point   = {v_load_max:.2f} mm  (incremental, load only)")
        print(f"  Δv theory (P_est)  = {v_theory:.2f} mm  (no real mass entered)")

    # ── Normal stress via Navier ───────────────────────────────────────────────
    x_all      = np.linspace(0, L_mm, 600)
    sigma_all  = np.abs(M_func(x_all)) * (h_mm / 2.0) / I_mm4   # [MPa]
    sigma_max  = float(np.max(sigma_all))
    x_peak     = float(x_all[np.argmax(sigma_all)])
    sigma_ref  = float(np.abs(M_func(np.array([a_mm]))[0]) * (h_mm/2.0) / I_mm4)
    print(f"  σ_max (Navier)     = {sigma_max:.4f} MPa  at x={x_peak:.0f} mm")
    print(f"  σ at ref. point    = {sigma_ref:.4f} MPa")
    print(f"{'='*54}\n")

    # ── Overlay image ─────────────────────────────────────────────────────────
    img_final = frame_loaded.copy()
    hi, wi    = img_final.shape[:2]

    # Draw loaded points (green) and unloaded points (cyan)
    for xo, yo in pts_loaded_px:
        cv2.circle(img_final, (int(xo), int(yo)), 3, (0,255,0), -1)
    for xo, yo in pts_unloaded_px:
        cv2.circle(img_final, (int(xo), int(yo)), 2, (0,200,255), -1)

    # Draw analytical fit curve (orange) — offset from unloaded reference
    x_fit       = np.linspace(0, L_mm, 800)
    dv_fit      = v_func(x_fit)
    v_ref_fit   = interpolate_reference(x_unl, v_unl, x_fit)
    v_total_fit = v_ref_fit + dv_fit   # absolute position in image

    direction = 1 if np.mean(np.array(pts_loaded_px)[:, 0]) > x_fix_px else -1
    xf_px = (x_fix_px + direction * x_fit * px_per_mm_x).astype(int)
    yf_px = (y_ref_adj + v_total_fit * px_per_mm_y).astype(int)
    for i in range(len(xf_px)-1):
        x1i, y1i = int(xf_px[i]),   int(yf_px[i])
        x2i, y2i = int(xf_px[i+1]), int(yf_px[i+1])
        if 0<=x1i<wi and 0<=y1i<hi and 0<=x2i<wi and 0<=y2i<hi:
            cv2.line(img_final, (x1i,y1i), (x2i,y2i), (0,140,255), 3)

    # Draw unloaded reference curve (cyan dashed approximation)
    xu_px = (x_fix_px + direction * x_unl * px_per_mm_x).astype(int)
    yu_px = (y_ref_adj + v_unl * px_per_mm_y).astype(int)
    for i in range(len(xu_px)-1):
        if 0<=xu_px[i]<wi and 0<=yu_px[i]<hi:
            cv2.line(img_final, (int(xu_px[i]),int(yu_px[i])),
                     (int(xu_px[i+1]),int(yu_px[i+1])), (0,200,255), 1)

    cv2.drawMarker(img_final, support_px, (255,0,255), cv2.MARKER_CROSS, 20, 2)
    # [display] Pink test label and colour legend removed: that information
    # (P, real mass, Delta v, colour key) now lives in the figure title, parameter
    # box, results box and the v(x) legend. The support cross is kept.

    d       = os.path.dirname(os.path.abspath(__file__))
    out_img = os.path.join(d, f"bda_overlay_test{test_id}.jpg")
    out_plt = os.path.join(d, f"bda_results_test{test_id}.png")
    cv2.imwrite(out_img, img_final)

    # [display only] Crop the report photo around the beam so it is larger and
    # better framed. Results above were computed on the full-resolution frame; this
    # only changes the embedded picture (full overlay JPG saved above is untouched).
    # Width is kept; height is trimmed but extended downwards so the hanging load
    # stays in view. Tune _pad_bot if you want more/less of the load.
    _ys  = [int(support_px[1])]
    _ys += [int(p[1]) for p in pts_loaded_px]
    _ys += [int(p[1]) for p in pts_unloaded_px]
    _ys += [int(v) for v in yf_px]
    _y0, _y1 = max(0, min(_ys)), min(hi, max(_ys))
    _span    = max(_y1 - _y0, 1)
    _pad_top = int(0.30 * _span) + 20
    _pad_bot = int(0.30 * _span) + int(0.20 * hi)   # extra room below for the load
    _r0, _r1 = max(0, _y0 - _pad_top), min(hi, _y1 + _pad_bot)
    img_show = img_final[_r0:_r1, :].copy()
    img_disp = cv2.resize(img_show, (int(img_show.shape[1] * scale),
                                     int(img_show.shape[0] * scale)))
    plot_results(x_load, dv, v_load, v_ref_interp,
                 v_func, M_func, V_func,
                 L_mm, P_N, a_mm, case,
                 img_overlay=img_disp, save_path=out_plt)

    result = {
        'test_id':      test_id,
        'case':         case,
        'a_mm':         a_mm,
        'params':       params,
        'mass_g_real':  mass_g,
        'P_N':          P_N,
        'mass_g_est':   mass_est,
        'delta_load':   v_load_max,
        'delta_self':   float(v_ref_interp[-1]),   # empirical, for reference
        'delta_total':  float(v_load[-1]),
        'delta_theory': v_theory,   # uses P_real if available, else P_N
        'overlay':      out_img,
        'plot':         out_plt,
    }

    save_results_excel(result, px_per_mm_x, d)
    return result




# ═══════════════════════════════════════════════════════════════════════════════
# STATIC-IMAGE TEST  (Mode B — no live camera)
# ═══════════════════════════════════════════════════════════════════════════════

def run_test_static(frame_unloaded, scale, grid,
                    px_per_mm_x, px_per_mm_y,
                    support_px, y_ref_adj,
                    test_id, case, params):
    """
    Identical workflow to run_test but uses file-dialog image loading
    instead of freeze_frame(cap, ...) for both the unloaded and loaded frames.
    The unloaded frame is already provided (captured during SETUP).
    Only the loaded frame is asked for here.
    """
    L_mm     = float(L_KNOWN_MM)
    x_fix_px = support_px[0]
    a_mm     = params.get('a_mm', L_mm)

    # ── Ask for real mass ─────────────────────────────────────────────────────
    mass_g = ask_mass_cv(test_id)
    print(f"  [TEST {test_id}] Real mass: "
          + (f"{mass_g:.1f} g" if mass_g else "not entered"))

    # ── Instructions ──────────────────────────────────────────────────────────
    img_w = np.zeros((300, 900, 3), dtype=np.uint8); img_w[:] = (25,25,25)
    cv2.putText(img_w, f"TEST {test_id}  [Mode B — static images]",
                (20,55), cv2.FONT_HERSHEY_SIMPLEX, 0.65, (0,255,255), 2)
    for i, (txt, col) in enumerate([
        ("The UNLOADED image was already loaded at setup.", (0,200,255)),
        ("You will now select the LOADED beam image.", (0,255,128)),
        ("Press SPACE to continue or ESC to cancel.", (200,200,0)),
    ]):
        cv2.putText(img_w, txt, (20, 120+i*50),
                    cv2.FONT_HERSHEY_SIMPLEX, 0.50, col, 1)
    cv2.imshow(WIN, img_w)
    while True:
        k = cv2.waitKey(50) & 0xFF
        if k == 27: print("  Test cancelled."); return None
        if k == ord(' '): break

    # ── Step 1: digitise unloaded beam ────────────────────────────────────────
    print(f"\n══ TEST {test_id} — STEP 1: DIGITISE UNLOADED BEAM ══════")
    pts_unloaded_px = digitize_unloaded(
        frame_unloaded, y_ref_adj, scale, px_per_mm_x, x_fix_px, grid)
    x_unl, v_unl = pts_to_mm(pts_unloaded_px, x_fix_px, y_ref_adj,
                              px_per_mm_x, px_per_mm_y, L_mm)

    # ── Step 2: load the loaded-beam image ────────────────────────────────────
    print(f"\n  [Mode B] Select the LOADED beam image for TEST {test_id}:")
    frame_loaded = load_static_image_cv()
    if frame_loaded is None:
        print("  No image selected — test cancelled."); return None

    # ── Step 3: digitise loaded beam ──────────────────────────────────────────
    print(f"\n══ TEST {test_id} — STEP 3: DIGITISE LOADED BEAM ══════")
    pts_loaded_px = digitize_deflection(
        frame_loaded, frame_unloaded, y_ref_adj,
        scale, px_per_mm_x, x_fix_px, grid)
    x_load, v_load = pts_to_mm(pts_loaded_px, x_fix_px, y_ref_adj,
                                px_per_mm_x, px_per_mm_y, L_mm)

    # ── Step 4: compute Δv ────────────────────────────────────────────────────
    v_ref_interp = interpolate_reference(x_unl, v_unl, x_load)
    dv           = v_load - v_ref_interp

    print(f"  Loaded tip v          = {v_load[-1]:.2f} mm")
    print(f"  Unloaded tip v (ref)  = {v_ref_interp[-1]:.2f} mm")
    print(f"  Incremental Δv tip    = {dv[-1]:.2f} mm")

    if np.max(np.abs(dv)) < 0.5:
        print("  [WARNING] Incremental deflection < 0.5 mm.")

    # ── Step 5: fit (identical to run_test) ───────────────────────────────────
    La = params.get('La', L_mm)
    Lb = params.get('Lb', 0.0)
    Lc = params.get('Lc', 0.0)

    if case == BeamCase.CANTILEVER_FREE_END:
        P_N, v_func, M_func, V_func = fit_cantilever_free_end(x_load, dv, L_mm)
    elif case == BeamCase.CANTILEVER_INTERMEDIATE:
        P_N, v_func, M_func, V_func = fit_cantilever_intermediate(x_load, dv, L_mm, a_mm)
    elif case == BeamCase.SIMPLY_SUPPORTED:
        P_N, v_func, M_func, V_func = fit_simply_supported(x_load, dv, L_mm, a_mm)
    elif case == BeamCase.OVERHANG_RIGHT_FREE_END:
        P_N, v_func, M_func, V_func = fit_overhang_right_free_end(x_load, dv, La, Lb)
    elif case == BeamCase.OVERHANG_LOAD_BETWEEN:
        P_N, v_func, M_func, V_func = fit_overhang_load_between(x_load, dv, La, Lb, Lc)
    elif case == BeamCase.OVERHANG_LOAD_IN_SPAN:
        P_N, v_func, M_func, V_func = fit_overhang_load_in_span(x_load, dv, La, Lb, Lc)

    if P_N is None: return None

    mass_est   = abs(P_N) / 9.81 * 1000
    x_ref_pt   = np.array([a_mm])
    v_load_max = float(v_func(x_ref_pt)[0])

    # Stress
    x_all     = np.linspace(0, L_mm, 600)
    sigma_all = np.abs(M_func(x_all)) * (h_mm/2.0) / I_mm4
    sigma_max = float(np.max(sigma_all))
    x_peak    = float(x_all[np.argmax(sigma_all)])

    print(f"\n{'='*54}")
    print(f"  TEST {test_id}  RESULTS  [Mode B]")
    print(f"  P (estimated)  = {P_N:.4f} N  (~{mass_est:.1f} g)")
    if mass_g:
        P_real = mass_g * 9.81 / 1000
        err_p  = (P_N - P_real) / P_real * 100
        print(f"  P (real)       = {P_real:.4f} N  ({mass_g:.1f} g)  "
              f"|  Error P = {err_p:+.1f}%")
    print(f"  σ_max (Navier) = {sigma_max:.4f} MPa  at x={x_peak:.0f} mm")
    print(f"{'='*54}\n")

    # ── Overlay and plot (reuse run_test logic) ───────────────────────────────
    img_final = frame_loaded.copy()
    hi, wi    = img_final.shape[:2]
    for xo, yo in pts_loaded_px:
        cv2.circle(img_final, (int(xo), int(yo)), 3, (0,255,0), -1)
    for xo, yo in pts_unloaded_px:
        cv2.circle(img_final, (int(xo), int(yo)), 2, (0,200,255), -1)

    x_fit       = np.linspace(0, L_mm, 800)
    dv_fit      = v_func(x_fit)
    v_ref_fit   = interpolate_reference(x_unl, v_unl, x_fit)
    v_total_fit = v_ref_fit + dv_fit
    direction   = 1 if np.mean(np.array(pts_loaded_px)[:, 0]) > x_fix_px else -1
    xf_px = (x_fix_px + direction * x_fit * px_per_mm_x).astype(int)
    yf_px = (y_ref_adj + v_total_fit * px_per_mm_y).astype(int)
    for i in range(len(xf_px)-1):
        x1i, y1i = int(xf_px[i]),   int(yf_px[i])
        x2i, y2i = int(xf_px[i+1]), int(yf_px[i+1])
        if 0<=x1i<wi and 0<=y1i<hi and 0<=x2i<wi and 0<=y2i<hi:
            cv2.line(img_final, (x1i,y1i), (x2i,y2i), (0,140,255), 3)

    # [display] Pink test label removed: P, real mass and Delta v now appear in the
    # figure title, parameter box and results box instead of over the photo.

    d       = os.path.dirname(os.path.abspath(__file__))
    out_img = os.path.join(d, f"bda_overlay_test{test_id}.jpg")
    out_plt = os.path.join(d, f"bda_results_test{test_id}.png")
    cv2.imwrite(out_img, img_final)

    # [display only] Crop the report photo around the beam so it is larger and
    # better framed. Results above were computed on the full-resolution frame; this
    # only changes the embedded picture (full overlay JPG saved above is untouched).
    # Width is kept; height is trimmed but extended downwards so the hanging load
    # stays in view. Tune _pad_bot if you want more/less of the load.
    _ys  = [int(support_px[1])]
    _ys += [int(p[1]) for p in pts_loaded_px]
    _ys += [int(p[1]) for p in pts_unloaded_px]
    _ys += [int(v) for v in yf_px]
    _y0, _y1 = max(0, min(_ys)), min(hi, max(_ys))
    _span    = max(_y1 - _y0, 1)
    _pad_top = int(0.30 * _span) + 20
    _pad_bot = int(0.30 * _span) + int(0.20 * hi)   # extra room below for the load
    _r0, _r1 = max(0, _y0 - _pad_top), min(hi, _y1 + _pad_bot)
    img_show = img_final[_r0:_r1, :].copy()
    img_disp = cv2.resize(img_show, (int(img_show.shape[1] * scale),
                                     int(img_show.shape[0] * scale)))
    plot_results(x_load, dv, v_load, v_ref_interp,
                 v_func, M_func, V_func,
                 L_mm, P_N, a_mm, case,
                 img_overlay=img_disp, save_path=out_plt)

    result = {
        'test_id':      test_id,
        'case':         case,
        'a_mm':         a_mm,
        'params':       params,
        'mass_g_real':  mass_g,
        'P_N':          P_N,
        'mass_g_est':   mass_est,
        'delta_load':   v_load_max,
        'delta_self':   float(v_ref_interp[-1]),
        'delta_total':  float(v_load[-1]),
        'delta_theory': 0.0,
        'overlay':      out_img,
        'plot':         out_plt,
    }
    save_results_excel(result, px_per_mm_x, d)
    return result


# ═══════════════════════════════════════════════════════════════════════════════
# EXCEL EXPORT
# ═══════════════════════════════════════════════════════════════════════════════

# Colour palette (openpyxl uses ARGB hex, no '#')
_BLUE_HEADER  = "FF003366"   # UPC dark blue  — header background
_BLUE_LIGHT   = "FFD6E4F0"   # light blue     — alternate data rows
_WHITE        = "FFFFFFFF"
_YELLOW_WARN  = "FFFFF0AA"   # soft yellow    — error > 10 %
_GREEN_GOOD   = "FFD6F0D6"   # soft green     — error ≤ 5 %
_ORANGE_MED   = "FFFFD6AA"   # soft orange    — error 5–10 %

_THIN = Side(style="thin", color="FF999999")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def _hdr_style(cell, text):
    """Apply bold white-on-blue header style."""
    cell.value = text
    cell.font      = Font(name="Arial", bold=True, color=_WHITE, size=10)
    cell.fill      = PatternFill("solid", start_color=_BLUE_HEADER)
    cell.alignment = Alignment(horizontal="center", vertical="center",
                               wrap_text=True)
    cell.border    = _BORDER


def _data_style(cell, value, fmt="General", bg=_WHITE):
    cell.value     = value
    cell.font      = Font(name="Arial", size=10)
    cell.fill      = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border    = _BORDER
    if fmt != "General":
        cell.number_format = fmt


def _error_colour(err_pct):
    """Return background colour based on absolute % error."""
    if err_pct is None:
        return _WHITE
    a = abs(err_pct)
    if a <= 5:   return _GREEN_GOOD
    if a <= 10:  return _ORANGE_MED
    return _YELLOW_WARN


def save_results_excel(result, px_per_mm, out_dir):
    """
    Append one test result to the session Excel workbook.

    File name: bda_session_YYYYMMDD.xlsx  (one file per calendar day,
    so multiple sessions on the same day accumulate in the same file).

    Sheet layout
    ------------
    Sheet "Session Log":
      One header row (written only when the sheet is new).
      One data row per test, appended each time this function is called.

    Sheet "Setup":
      Beam and calibration parameters — written once per file creation,
      updated if the file already exists.

    Sheet "x-v data  TESTn":
      Full digitised + fitted deflection table for test n, plus
      M(x) and V(x) columns computed from the fitted analytical solution.
    """
    if not EXCEL_AVAILABLE:
        print("[EXCEL] openpyxl not available — skipping export.")
        return None

    # ── File path ─────────────────────────────────────────────────────────────
    today    = datetime.date.today().strftime("%Y%m%d")
    xlsx_path = os.path.join(out_dir, f"bda_session_{today}.xlsx")

    case_str = {
        BeamCase.CANTILEVER_FREE_END:     "Cantilever — free end",
        BeamCase.CANTILEVER_INTERMEDIATE: "Cantilever — intermediate",
        BeamCase.SIMPLY_SUPPORTED:        "Simply supported",
        BeamCase.OVERHANG_RIGHT_FREE_END: "Overhanging — load at free end",
        BeamCase.OVERHANG_LOAD_BETWEEN:   "Overhanging — load between supports",
        BeamCase.OVERHANG_LOAD_IN_SPAN:   "Overhanging — load in overhang",
    }[result['case']]

    # ── Load or create workbook ────────────────────────────────────────────────
    if os.path.exists(xlsx_path):
        wb = load_workbook(xlsx_path)
    else:
        wb = Workbook()
        # Remove default blank sheet
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

    # ══════════════════════════════════════════════════════════════════════════
    # Sheet 1 — Setup (beam properties + calibration)
    # ══════════════════════════════════════════════════════════════════════════
    if "Setup" not in wb.sheetnames:
        ws_setup = wb.create_sheet("Setup", 0)

        ws_setup.column_dimensions["A"].width = 34
        ws_setup.column_dimensions["B"].width = 18
        ws_setup.column_dimensions["C"].width = 14

        # Title
        ws_setup.merge_cells("A1:C1")
        t = ws_setup["A1"]
        t.value     = "Beam Deflection Analysis — Beam & Calibration Setup"
        t.font      = Font(name="Arial", bold=True, size=12, color=_WHITE)
        t.fill      = PatternFill("solid", start_color=_BLUE_HEADER)
        t.alignment = Alignment(horizontal="center", vertical="center")

        rows = [
            ("BEAM CROSS-SECTION", "", ""),
            ("Width  b", f"{b_mm}", "mm"),
            ("Height  h", f"{h_mm}", "mm"),
            ("Second moment of area  I", f"{I_mm4:.4f}", "mm⁴"),
            ("", "", ""),
            ("MATERIAL PROPERTIES", "", ""),
            ("Young's modulus  E", f"{E_MPa:.0f}", "MPa"),
            ("Density  ρ", f"{rho_kgm3:.0f}", "kg/m³"),
            ("Flexural rigidity  EI", f"{EI:.2f}", "N·mm²"),
            ("Self-weight per unit length  w", f"{w_Nmm*1000:.4f}", "mN/mm"),
            ("", "", ""),
            ("EXPERIMENTAL SETUP", "", ""),
            ("Beam span  L", f"{L_KNOWN_MM:.1f}", "mm"),
            ("Calibration  px/mm (isotropic)", f"{px_per_mm:.4f}", "px/mm"),
            ("Self-weight handling", "Empirical (two-digitisation)", ""),
            ("", "", ""),
            ("Generated", datetime.datetime.now().strftime("%Y-%m-%d  %H:%M:%S"), ""),
        ]

        section_keys = {"BEAM CROSS-SECTION", "MATERIAL PROPERTIES",
                        "EXPERIMENTAL SETUP"}

        for i, (label, value, unit) in enumerate(rows, start=2):
            ws_setup[f"A{i}"] = label
            ws_setup[f"B{i}"] = value
            ws_setup[f"C{i}"] = unit
            if label in section_keys:
                for col in ("A", "B", "C"):
                    ws_setup[f"{col}{i}"].font = Font(
                        name="Arial", bold=True, size=10, color=_WHITE)
                    ws_setup[f"{col}{i}"].fill = PatternFill(
                        "solid", start_color="FF336699")
            else:
                for col in ("A", "B", "C"):
                    ws_setup[f"{col}{i}"].font = Font(name="Arial", size=10)
                    ws_setup[f"{col}{i}"].alignment = Alignment(
                        horizontal="left" if col == "A" else "center")
    else:
        ws_setup = wb["Setup"]

    # ══════════════════════════════════════════════════════════════════════════
    # Sheet 2 — Session Log (one row per test)
    # ══════════════════════════════════════════════════════════════════════════
    LOG_HEADERS = [
        "Test #", "Timestamp", "Beam case", "Load pos. a [mm]",
        "Mass real [g]", "P real [N]",
        "P estimated [N]", "Mass est. [g]", "Error P [%]",
        "δ_load [mm]", "δ_self [mm]", "δ_total [mm]",
        "Overlay file", "Plot file",
    ]

    if "Session Log" not in wb.sheetnames:
        ws_log = wb.create_sheet("Session Log", 1)
        # Title row
        ws_log.merge_cells(f"A1:{get_column_letter(len(LOG_HEADERS))}1")
        tc = ws_log["A1"]
        tc.value     = "Beam Deflection Analysis — Session Log"
        tc.font      = Font(name="Arial", bold=True, size=12, color=_WHITE)
        tc.fill      = PatternFill("solid", start_color=_BLUE_HEADER)
        tc.alignment = Alignment(horizontal="center", vertical="center")
        ws_log.row_dimensions[1].height = 22

        # Header row
        for col_i, h in enumerate(LOG_HEADERS, start=1):
            _hdr_style(ws_log.cell(row=2, column=col_i), h)
        ws_log.row_dimensions[2].height = 32
        ws_log.freeze_panes = "A3"

        # Column widths
        col_widths = [7, 18, 26, 14, 13, 12, 15, 13, 11,
                      12, 12, 13, 34, 34]
        for ci, w in enumerate(col_widths, start=1):
            ws_log.column_dimensions[get_column_letter(ci)].width = w
    else:
        ws_log = wb["Session Log"]

    # Next data row
    next_row = ws_log.max_row + 1
    bg = _BLUE_LIGHT if next_row % 2 == 0 else _WHITE

    # Compute derived values for this test
    P_real   = result['mass_g_real'] * 9.81 / 1000 if result['mass_g_real'] else None
    err_p    = ((result['P_N'] - P_real) / P_real * 100) if P_real else None
    err_d    = ((result['delta_total'] - result['delta_theory']) /
                result['delta_theory'] * 100) if result['delta_theory'] else None
    err_bg   = _error_colour(err_p)

    row_vals = [
        (result['test_id'],                    "General", bg),
        (datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"), "General", bg),
        (case_str,                             "General", bg),
        (result['a_mm'],                       "0.0",     bg),
        (result['mass_g_real'] or "",          "0.0",     bg),
        (P_real or "",                         "0.0000",  bg),
        (result['P_N'],                        "0.0000",  bg),
        (result['mass_g_est'],                 "0.0",     bg),
        (err_p if err_p is not None else "",   "+0.0;-0.0;-", err_bg),
        (result['delta_load'],                 "0.00",    bg),
        (result['delta_self'],                 "0.00",    bg),
        (result['delta_total'],                "0.00",    bg),
        (os.path.basename(result['overlay']),  "General", bg),
        (os.path.basename(result['plot']),     "General", bg),
    ]

    for ci, (val, fmt, bg_c) in enumerate(row_vals, start=1):
        _data_style(ws_log.cell(row=next_row, column=ci), val, fmt, bg_c)

    # ══════════════════════════════════════════════════════════════════════════
    # Sheet 3 — Detailed x-v-M-V table for this test
    # ══════════════════════════════════════════════════════════════════════════
    # We need v_func, M_func, V_func — they are not in result dict,
    # so we reconstruct using the stored P_N and analytical formulas.
    P   = result['P_N']
    L   = L_KNOWN_MM
    a   = result['a_mm']
    pr  = result.get('params', {})
    La  = pr.get('La', L)
    Lb  = pr.get('Lb', 0.0)
    Lc  = pr.get('Lc', 0.0)

    if result['case'] == BeamCase.CANTILEVER_FREE_END:
        def _v(x): return P/(6*EI)*(3*L*x**2-x**3)
        def _M(x): return P*(L-np.asarray(x))
        def _V(x): return np.full_like(np.asarray(x,float), P)
    elif result['case'] == BeamCase.CANTILEVER_INTERMEDIATE:
        def _v(x):
            x=np.asarray(x,float); r=np.zeros_like(x)
            m1=x<=a; m2=~m1
            r[m1]=P/(6*EI)*(3*a*x[m1]**2-x[m1]**3)
            r[m2]=P*a**2/(6*EI)*(3*x[m2]-a)
            return r
        def _M(x):
            x=np.asarray(x,float); r=np.zeros_like(x)
            r[x<=a]=P*(a-x[x<=a]); return r
        def _V(x):
            x=np.asarray(x,float); r=np.zeros_like(x)
            r[x<=a]=P; return r
    elif result['case'] == BeamCase.SIMPLY_SUPPORTED:
        b_ss=L-a; Ra=P*b_ss/L
        def _v(x):
            x=np.asarray(x,float); r=np.zeros_like(x)
            m1=x<=a; m2=~m1
            r[m1]=P*b_ss/(6*EI*L)*(L**2*x[m1]-b_ss**2*x[m1]-x[m1]**3)
            r[m2]=P*a*(L-x[m2])/(6*EI*L)*(2*L*x[m2]-x[m2]**2-a**2)
            return r
        def _M(x):
            x=np.asarray(x,float); r=np.zeros_like(x)
            m1=x<=a; m2=~m1
            r[m1]=Ra*x[m1]; r[m2]=Ra*x[m2]-P*(x[m2]-a); return r
        def _V(x):
            x=np.asarray(x,float); r=np.zeros_like(x)
            r[x<=a]=Ra; r[~(x<=a)]=Ra-P; return r
    elif result['case'] == BeamCase.OVERHANG_RIGHT_FREE_END:
        RA = -P*Lb/La;  RB = P*(La+Lb)/La
        def _v(x):
            x=np.asarray(x,float); r=np.zeros_like(x)
            m1=x<=La; m2=~m1
            r[m1]=-P*Lb*x[m1]*(x[m1]**2-La**2)/(6*EI*La)
            xi=x[m2]; r[m2]=P*(xi-La)*(xi**2-2*(La+Lb)*xi+La**2+La*Lb)/(6*EI)
            return r
        def _M(x):
            x=np.asarray(x,float); r=np.zeros_like(x)
            m1=x<=La; m2=~m1
            r[m1]=-(P*Lb/La)*x[m1]; r[m2]=-P*(La+Lb-x[m2]); return r
        def _V(x):
            x=np.asarray(x,float); r=np.zeros_like(x)
            r[x<=La]=RA; r[x>La]=P; return r
    elif result['case'] == BeamCase.OVERHANG_LOAD_BETWEEN:
        Sab=La+Lb; RA=P*Lb/Sab; RC=P*La/Sab
        def _v(x):
            x=np.asarray(x,float); r=np.zeros_like(x)
            m1=x<=La; m2=(x>La)&(x<=Sab); m3=x>Sab
            r[m1]=P*Lb*x[m1]*(x[m1]**2-La**2-2*La*Lb)/(6*EI*Sab)
            xi=x[m2]; r[m2]=-P*La*(xi-Sab)*(xi**2-2*Sab*xi+La**2)/(6*EI*Sab)
            r[m3]=P*La*Lb*(2*La+Lb)*(x[m3]-Sab)/(6*EI*Sab); return r
        def _M(x):
            x=np.asarray(x,float); r=np.zeros_like(x)
            r[x<=La]=RA*x[x<=La]
            m2=(x>La)&(x<=Sab); r[m2]=P*La/Sab*(Sab-x[m2]); return r
        def _V(x):
            x=np.asarray(x,float); r=np.zeros_like(x)
            r[x<=La]=RA; r[(x>La)&(x<=Sab)]=RA-P; return r
    elif result['case'] == BeamCase.OVERHANG_LOAD_IN_SPAN:
        Sab=La+Lb; RA=-P*Lb/La; RB=P*Sab/La
        def _v(x):
            x=np.asarray(x,float); r=np.zeros_like(x)
            m1=x<=La; m2=(x>La)&(x<=Sab); m3=x>Sab
            r[m1]=-P*Lb*x[m1]*(x[m1]-La)*(x[m1]+La)/(6*EI*La)
            xi=x[m2]; r[m2]=P*(xi-La)*(xi**2-(2*La+3*Lb)*xi+La**2+La*Lb)/(6*EI)
            r[m3]=-P*Lb*(-2*La**2-3*La*Lb+2*La*x[m3]-Lb**2+3*Lb*x[m3])/(6*EI); return r
        def _M(x):
            x=np.asarray(x,float); r=np.zeros_like(x)
            r[x<=La]=RA*x[x<=La]
            m2=(x>La)&(x<=Sab); r[m2]=P*(x[m2]-Sab); return r
        def _V(x):
            x=np.asarray(x,float); r=np.zeros_like(x)
            r[x<=La]=RA; r[(x>La)&(x<=Sab)]=RA+RB; return r

    sheet_name = f"Data TEST{result['test_id']}"
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws_data = wb.create_sheet(sheet_name)

    # Title
    DATA_COLS = ["x [mm]", "Δv_fit [mm]", "M(x) [N·mm]", "V(x) [N]"]
    ws_data.merge_cells(f"A1:{get_column_letter(len(DATA_COLS))}1")
    tc2 = ws_data["A1"]
    tc2.value     = (f"Test {result['test_id']} — {case_str}  |  "
                     f"P = {P:.4f} N  |  L = {L:.0f} mm  |  "
                     f"EI = {EI:.0f} N·mm²")
    tc2.font      = Font(name="Arial", bold=True, size=11, color=_WHITE)
    tc2.fill      = PatternFill("solid", start_color=_BLUE_HEADER)
    tc2.alignment = Alignment(horizontal="center", vertical="center")
    ws_data.row_dimensions[1].height = 20

    for ci, h in enumerate(DATA_COLS, start=1):
        _hdr_style(ws_data.cell(row=2, column=ci), h)
    ws_data.freeze_panes = "A3"

    x_arr      = np.linspace(0, L, 200)
    v_fit_arr  = _v(x_arr)    # incremental deflection Δv (load only)
    M_arr      = _M(x_arr)
    V_arr      = _V(x_arr)

    rms_vals = []
    for ri, xi in enumerate(x_arr, start=3):
        idx = ri - 3
        bg_r = _BLUE_LIGHT if ri % 2 == 0 else _WHITE
        row_data = [
            (float(xi),              "0.00"),
            (float(v_fit_arr[idx]),  "0.000"),
            (float(M_arr[idx]),      "0.0"),
            (float(V_arr[idx]),      "0.0000"),
        ]
        for ci, (val, fmt) in enumerate(row_data, start=1):
            _data_style(ws_data.cell(row=ri, column=ci), val, fmt, bg_r)

    for ci in range(1, len(DATA_COLS)+1):
        ws_data.column_dimensions[get_column_letter(ci)].width = 16

    # ── Save ──────────────────────────────────────────────────────────────────
    wb.save(xlsx_path)
    print(f"[EXCEL] Results appended → {os.path.basename(xlsx_path)}")
    return xlsx_path


# ═══════════════════════════════════════════════════════════════════════════════
# SESSION SUMMARY
# ═══════════════════════════════════════════════════════════════════════════════

def print_summary(results, L_mm):
    print(f"\n{'═'*80}")
    print(f"  SESSION SUMMARY  —  {len(results)} test(s)")
    print(f"{'═'*80}")
    print(f"  {'Test':>4}  {'Mass real':>10}  {'P est (N)':>10}  "
          f"{'Mass est (g)':>13}  {'δ_load':>8}  {'δ_total':>10}  "
          f"{'δ_theory':>10}  {'Error P':>8}")
    print(f"  {'-'*74}")

    with_real = []
    for r in results:
        err_str  = ""
        theo_str = f"{r['delta_theory']:.1f}"
        if r['mass_g_real']:
            P_real  = r['mass_g_real'] * 9.81 / 1000
            err_str = f"{(r['P_N']-P_real)/P_real*100:+.1f}%"
            with_real.append(r)
        mass_str = f"{r['mass_g_real']:.0f} g" if r['mass_g_real'] else "—"
        print(f"  {r['test_id']:>4}  {mass_str:>10}  {r['P_N']:>10.4f}  "
              f"{r['mass_g_est']:>13.1f}  {r['delta_load']:>8.1f}  "
              f"{r['delta_total']:>10.1f}  {theo_str:>10}  {err_str:>8}")

    print(f"{'═'*80}\n")

    if len(with_real) >= 2:
        errors = [(r['P_N'] - r['mass_g_real']*9.81/1000) /
                  (r['mass_g_real']*9.81/1000) * 100
                  for r in with_real]
        print(f"  Mean error in P: {np.mean(errors):+.1f}%  "
              f"(std = {np.std(errors):.1f}%)")
        print(f"  If this is consistently non-zero, check:")
        print(f"    1. Calibration: are the two clicked points exactly L_KNOWN_MM apart?")
        print(f"    2. Camera tilt: consider optional vertical calibration (menu R -> recal).")
        print(f"    3. EI value: verify b, h, E with independent measurement.")


# ═══════════════════════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════════════════════

def main():
    print("\n" + "═"*60)
    print("  BEAM DEFLECTION & STRESS ANALYSIS")
    print(f"  EI = {EI:.0f} N·mm²   L = {L_KNOWN_MM:.0f} mm")
    print("═"*60 + "\n")

    cv2.namedWindow(WIN, cv2.WINDOW_NORMAL)
    cv2.resizeWindow(WIN, 1280, 720)
    grid = Grid(1280, 720)

    # ── Mode selection: A = live camera, B = static image ────────────────────
    mode = select_mode_cv()

    if mode == 'live':
        cap          = open_camera(CAMERA_INDEX)
        scale, _     = get_scale(cap)

        def acquire_frame(msg):
            return freeze_frame(cap, scale, msg, grid)

        def release():
            cap.release()

    else:  # static image mode
        cap   = None
        scale = 1.0

        # Load the image once; use it as both "unloaded" and provide a second
        # load for the loaded state during the test.

        def release():
            pass   # no camera to release

    # ── Select beam case (once per session) ───────────────────────────────────
    case, params = select_beam_case_cv()
    a_mm = params.get('a_mm', L_KNOWN_MM)

    case_labels = {
        BeamCase.CANTILEVER_FREE_END:     "Cantilever — load at free end",
        BeamCase.CANTILEVER_INTERMEDIATE: f"Cantilever — load at a={a_mm:.0f} mm",
        BeamCase.SIMPLY_SUPPORTED:        f"Simply supported — load at a={a_mm:.0f} mm",
        BeamCase.OVERHANG_RIGHT_FREE_END: f"Overhanging — load at right free end"
                                          f"  (La={params.get('La',0):.0f} Lb={params.get('Lb',0):.0f} mm)",
        BeamCase.OVERHANG_LOAD_BETWEEN:   f"Overhanging — load between supports"
                                          f"  (La={params.get('La',0):.0f} Lb={params.get('Lb',0):.0f}"
                                          f" Lc={params.get('Lc',0):.0f} mm)",
        BeamCase.OVERHANG_LOAD_IN_SPAN:   f"Overhanging — load in right overhang"
                                          f"  (La={params.get('La',0):.0f} Lb={params.get('Lb',0):.0f}"
                                          f" Lc={params.get('Lc',0):.0f} mm)",
    }
    case_label = case_labels[case]

    # ── SETUP (once per session) ───────────────────────────────────────────────
    print("═"*56)
    print(" SETUP — unloaded beam + calibration")
    print("═"*56)

    # 1. Unloaded frame
    if mode == 'live':
        frame_unloaded = freeze_frame(cap, scale,
                                      "Beam WITHOUT LOAD -> SPACE to capture", grid)
    else:
        # In static mode, ask for the unloaded image explicitly
        print("\n  [Mode B] Select the UNLOADED beam image:")
        frame_unloaded = load_static_image_cv()
        if frame_unloaded is None:
            print("  No image — exiting."); cv2.destroyAllWindows(); return
        h_s, w_s = frame_unloaded.shape[:2]
        scale    = min(1.0, 1280/w_s, 720/h_s)

    # 2. Isotropic calibration
    px_per_mm_x = calibrate(frame_unloaded, scale, grid)
    px_per_mm_y = px_per_mm_x

    # 3. Reference support
    print("\n  RIGHT-CLICK on the reference support")
    support_px = zoom_one_right_click(
        frame_unloaded, scale,
        "RIGHT-CLICK: reference support  |  +/-: zoom", grid)
    print(f"  Support: x={support_px[0]} px  y={support_px[1]} px")
    grid.set_origin(int(support_px[0]*scale), int(support_px[1]*scale))

    # 4. Neutral axis
    print("\n" + "═"*56)
    print(" SETUP — align neutral axis")
    print("═"*56)
    y_ref_adj = adjust_neutral_axis(
        frame_unloaded, frame_unloaded, support_px[1], scale, grid)
    grid.set_origin(int(support_px[0]*scale), int(y_ref_adj*scale))

    print(f"\n  Setup complete:")
    print(f"    px_per_mm = {px_per_mm_x:.4f} px/mm  (isotropic)")
    print(f"    Support pixel: ({support_px[0]}, {support_px[1]})")
    print(f"    Neutral axis y = {y_ref_adj:.1f} px")

    # ── Session loop ───────────────────────────────────────────────────────────
    results = []
    test_id = 1

    while True:
        cmd = show_menu_cv(test_id, results, case_label)

        if cmd == 'quit':
            break

        if cmd == 'recal':
            print("\n  Recalibrating...")
            if mode == 'live':
                frame_unloaded = freeze_frame(
                    cap, scale, "Beam WITHOUT LOAD -> SPACE", grid)
            else:
                _new_unl = load_static_image_cv()
                if _new_unl is not None:
                    frame_unloaded = _new_unl
            px_per_mm_x = calibrate(frame_unloaded, scale, grid)
            px_per_mm_y = px_per_mm_x

            img_q = np.zeros((200, 800, 3), dtype=np.uint8); img_q[:] = (25,25,25)
            cv2.putText(img_q,
                "Add separate VERTICAL calibration? (Y = yes, N = skip)",
                (20,90), cv2.FONT_HERSHEY_SIMPLEX, 0.60, (0,255,255), 2)
            cv2.imshow(WIN, img_q)
            while True:
                k = cv2.waitKey(50) & 0xFF
                if k in (ord('y'), ord('Y')):
                    px_per_mm_y = calibrate_vertical(
                        frame_unloaded, scale, grid, px_per_mm_x)
                    break
                if k in (ord('n'), ord('N'), 27): break

            support_px = zoom_one_right_click(
                frame_unloaded, scale, "RIGHT-CLICK: reference support", grid)
            grid.set_origin(int(support_px[0]*scale), int(support_px[1]*scale))
            y_ref_adj = adjust_neutral_axis(
                frame_unloaded, frame_unloaded, support_px[1], scale, grid)
            grid.set_origin(int(support_px[0]*scale), int(y_ref_adj*scale))
            print("  Recalibration complete.")
            continue

        # cmd == 'test'
        if mode == 'static':
            result = run_test_static(frame_unloaded, scale, grid,
                                     px_per_mm_x, px_per_mm_y,
                                     support_px, y_ref_adj,
                                     test_id, case, params)
        else:
            result = run_test(cap, scale, grid, frame_unloaded,
                              px_per_mm_x, px_per_mm_y,
                              support_px, y_ref_adj,
                              test_id, case, params)
        if result:
            results.append(result)
            test_id += 1

    release()
    cv2.destroyAllWindows()

    if results:
        print_summary(results, float(L_KNOWN_MM))
        print("  Saved files:")
        for r in results:
            print(f"    TEST {r['test_id']}: {os.path.basename(r['plot'])}")
    else:
        print("\n  No tests were completed.")


if __name__ == "__main__":
    main()
